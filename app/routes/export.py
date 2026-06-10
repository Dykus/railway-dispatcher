# app/routes/export.py
# -*- coding: utf-8 -*-
"""
Маршруты для экспорта данных в Excel.
"""

from flask import Blueprint, send_file, flash, redirect, url_for, request, g
import io
import pandas as pd
from datetime import datetime
import sys
import os
from openpyxl.styles import Font, Alignment

sys.path.append(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
from app.models import get_conn, calculate_overstay, calculate_current_overstay, get_setting
from app.utils import get_moscow_now

export_bp = Blueprint('export', __name__)


def apply_excel_styling(writer, sheet_name, has_notes=False):
    from openpyxl.styles import Font, Alignment
    worksheet = writer.sheets[sheet_name]
    for column in worksheet.columns:
        max_length = 0
        col_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[col_letter].width = adjusted_width
    worksheet.auto_filter.ref = worksheet.dimensions
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            if has_notes and cell.column_letter in ('F', 'G', 'H', 'D', 'E'):
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')


@export_bp.route('/export_excel')
def export_excel():
    station_id = g.get('station_id', 1)
    conn = get_conn()
    df = pd.read_sql_query("""
        SELECT 
            wv.wagon_number as "Номер вагона", 
            wv.owner as "Транспортная компания", 
            wv.organization as "Организация", 
            t.name as "Путь", 
            wv.arrival_time as "Время прибытия", 
            wv.departure_time as "Глобальный срок" 
        FROM wagon_visits wv 
        JOIN tracks t ON wv.track_id = t.id 
        WHERE wv.status != 'departed' AND wv.is_archived = 0 AND wv.station_id = ?
        ORDER BY wv.wagon_number
    """, conn, params=(station_id,))
    conn.close()

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Отчет', index=False)
        apply_excel_styling(writer, 'Отчет')

    output.seek(0)
    return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=f"Report_{get_moscow_now().strftime('%Y%m%d_%H%M%S')}.xlsx")


@export_bp.route('/export_history_excel')
def export_history_excel():
    station_id = g.get('station_id', 1)
    conn = get_conn()
    df = pd.read_sql_query("""
        SELECT 
            m.wagon_number as "Номер вагона", 
            m.action_type as "Тип действия", 
            m.from_track as "Откуда", 
            m.to_track as "Куда", 
            wv.owner as "Транспортная компания",
            wv.organization as "Организация",
            m.note as "Примечание", 
            m.timestamp as "Время" 
        FROM movement_history m
        LEFT JOIN wagon_visits wv ON m.wagon_number = wv.wagon_number AND wv.is_archived = 0 AND wv.station_id = ?
        WHERE m.station_id = ?
        ORDER BY m.timestamp DESC
    """, conn, params=(station_id, station_id))
    conn.close()

    action_map = {'added': 'Добавлен', 'moved': 'Перемещен', 'departed': 'Убыл', 'edit': 'Изменён'}
    df['Тип действия'] = df['Тип действия'].map(action_map).fillna(df['Тип действия'])

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='История', index=False)
        apply_excel_styling(writer, 'История', has_notes=True)

    output.seek(0)
    return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=f"History_{get_moscow_now().strftime('%Y%m%d_%H%M%S')}.xlsx")


@export_bp.route('/export_archive_excel')
def export_archive_excel():
    station_id = g.get('station_id', 1)
    conn = get_conn()
    filter_type = request.args.get('filter_type', 'all')
    year = request.args.get('year', '')
    month = request.args.get('month', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')

    sql_summary = """
        SELECT 
            wv.wagon_number as "Номер вагона", 
            wv.owner as "Транспортная компания", 
            wv.organization as "Организация", 
            wv.arrival_time as "Время прибытия",
            wv.departure_time as "Глобальный срок",
            (SELECT MAX(timestamp) FROM archived_history WHERE visit_id = wv.id) as "Фактическое убытие",
            wv.id as visit_id
        FROM wagon_visits wv
        WHERE wv.is_archived = 1 AND wv.station_id = ?
    """
    sql_details = """
        SELECT 
            a.wagon_number as "Номер вагона", 
            a.action_type as "Тип действия", 
            a.from_track as "Откуда", 
            a.to_track as "Куда", 
            a.note as "Примечание", 
            a.timestamp as "Время"
        FROM archived_history a
        WHERE a.station_id = ?
    """

    params_summary = [station_id]
    params_details = [station_id]

    if filter_type == 'year' and year:
        sql_summary += " AND substr(wv.arrival_time, 1, 4) = ?"
        params_summary.append(year)
        sql_details += " AND a.visit_id IN (SELECT id FROM wagon_visits WHERE is_archived = 1 AND station_id = ? AND substr(arrival_time, 1, 4) = ?)"
        params_details.extend([station_id, year])
    elif filter_type == 'month' and year and month:
        period = f"{year}-{month}"
        sql_summary += " AND substr(wv.arrival_time, 1, 7) = ?"
        params_summary.append(period)
        sql_details += " AND a.visit_id IN (SELECT id FROM wagon_visits WHERE is_archived = 1 AND station_id = ? AND substr(arrival_time, 1, 7) = ?)"
        params_details.extend([station_id, period])
    elif filter_type == 'period' and date_from and date_to:
        sql_summary += " AND date(wv.arrival_time) BETWEEN ? AND ?"
        params_summary.extend([date_from, date_to])
        sql_details += " AND a.visit_id IN (SELECT id FROM wagon_visits WHERE is_archived = 1 AND station_id = ? AND date(arrival_time) BETWEEN ? AND ?)"
        params_details.extend([station_id, date_from, date_to])

    sql_summary += " ORDER BY wv.wagon_number"
    sql_details += " ORDER BY a.wagon_number, a.timestamp ASC"

    df_summary = pd.read_sql_query(sql_summary, conn, params=params_summary)
    df_details = pd.read_sql_query(sql_details, conn, params=params_details)
    conn.close()

    action_map = {'added': 'Добавлен', 'moved': 'Перемещен', 'departed': 'Убыл', 'edit': 'Изменён'}
    df_details['Тип действия'] = df_details['Тип действия'].map(action_map).fillna(df_details['Тип действия'])

    try:
        df_summary['Время прибытия'] = pd.to_datetime(df_summary['Время прибытия']).dt.strftime('%d-%m-%Y %H:%M:%S')
        df_summary['Глобальный срок'] = pd.to_datetime(df_summary['Глобальный срок']).dt.strftime('%d-%m-%Y %H:%M:%S')
        df_summary['Фактическое убытие'] = pd.to_datetime(df_summary['Фактическое убытие']).dt.strftime('%d-%m-%Y %H:%M:%S')
    except:
        pass
    try:
        df_details['Время'] = pd.to_datetime(df_details['Время']).dt.strftime('%d-%m-%Y %H:%M:%S')
    except:
        pass

    overstays = []
    amounts = []
    for _, row in df_summary.iterrows():
        vid = row['visit_id']
        over, amt = calculate_overstay(vid)
        overstays.append(over if over > 0 else '')
        amounts.append(amt if amt > 0 else None)   # числовой формат
    df_summary['Перепростой, сут'] = overstays
    df_summary['Сумма, руб'] = amounts
    df_summary.drop(columns=['visit_id'], inplace=True, errors='ignore')

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_summary.to_excel(writer, sheet_name='Сводка', index=False)
        apply_excel_styling(writer, 'Сводка')
        df_details.to_excel(writer, sheet_name='Детализация', index=False)
        apply_excel_styling(writer, 'Детализация', has_notes=True)

    output.seek(0)
    return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=f"Archive_{get_moscow_now().strftime('%Y%m%d_%H%M%S')}.xlsx")


@export_bp.route('/export_wagon_history/<wagon_number>')
def export_wagon_history(wagon_number):
    station_id = g.get('station_id', 1)
    conn = get_conn()
    df = pd.read_sql_query("""
        SELECT action_type as "Тип действия", from_track as "Откуда", to_track as "Куда", note as "Примечание", timestamp as "Время" 
        FROM movement_history 
        WHERE wagon_number = ? AND station_id = ?
        ORDER BY timestamp ASC
    """, conn, params=(wagon_number, station_id))
    conn.close()

    if df.empty:
        flash(f"Нет данных по вагону {wagon_number}", 'error')
        return redirect(url_for('history.history_page', station_id=station_id))

    action_map = {'added': 'Добавлен', 'moved': 'Перемещен', 'departed': 'Убыл', 'edit': 'Изменён'}
    df['Тип действия'] = df['Тип действия'].map(action_map).fillna(df['Тип действия'])
    try:
        df['Время'] = pd.to_datetime(df['Время']).dt.strftime('%d-%m-%Y %H:%M:%S')
    except:
        pass

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=f'История {wagon_number}', index=False)
        apply_excel_styling(writer, f'История {wagon_number}', has_notes=True)

    output.seek(0)
    return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=f"History_{wagon_number}_{get_moscow_now().strftime('%Y%m%d_%H%M%S')}.xlsx")


@export_bp.route('/export_wagon_archive/<wagon_number>')
def export_wagon_archive(wagon_number):
    station_id = g.get('station_id', 1)
    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT id FROM wagon_visits WHERE wagon_number = ? AND is_archived = 1 AND station_id = ? ORDER BY id DESC LIMIT 1", (wagon_number, station_id))
    row = c.fetchone()
    if not row:
        flash(f"Нет архивных данных по вагону {wagon_number}", 'error')
        conn.close()
        return redirect(url_for('history.archive_page', station_id=station_id))
    visit_id = row[0]

    df = pd.read_sql_query("""
        SELECT action_type as "Тип действия", from_track as "Откуда", to_track as "Куда", note as "Примечание", timestamp as "Время" 
        FROM archived_history 
        WHERE visit_id = ? AND station_id = ?
        ORDER BY timestamp ASC
    """, conn, params=(visit_id, station_id))
    conn.close()

    if df.empty:
        flash(f"Нет данных по вагону {wagon_number} в архиве", 'error')
        return redirect(url_for('history.archive_page', station_id=station_id))

    action_map = {'added': 'Добавлен', 'moved': 'Перемещен', 'departed': 'Убыл', 'edit': 'Изменён'}
    df['Тип действия'] = df['Тип действия'].map(action_map).fillna(df['Тип действия'])
    try:
        df['Время'] = pd.to_datetime(df['Время']).dt.strftime('%d-%m-%Y %H:%M:%S')
    except:
        pass

    overstay, amount = calculate_overstay(visit_id)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=f'Архив {wagon_number}', index=False)
        worksheet = writer.sheets[f'Архив {wagon_number}']
        
        # Определяем последнюю строку с данными (заголовок в строке 1)
        last_data_row = len(df) + 1
        
        # Пустая строка для отступа
        empty_row = last_data_row + 1
        worksheet.cell(row=empty_row, column=1, value='')
        
        # Строка с перепростоем в колонках F и G (6 и 7)
        row_overstay = empty_row + 1
        cell_label = worksheet.cell(row=row_overstay, column=6, value='Перепростой, сут:')
        cell_value = worksheet.cell(row=row_overstay, column=7, value=str(overstay) if overstay > 0 else '')
        
        # Строка с суммой
        row_sum = row_overstay + 1
        cell_label2 = worksheet.cell(row=row_sum, column=6, value='Сумма, руб:')
        cell_value2 = worksheet.cell(row=row_sum, column=7, value=f"{amount:.2f}" if amount > 0 else '')
        
        # Жирный шрифт для этих ячеек
        bold_font = Font(bold=True)
        for cell in [cell_label, cell_value, cell_label2, cell_value2]:
            cell.font = bold_font
        
        apply_excel_styling(writer, f'Архив {wagon_number}', has_notes=True)
    
    output.seek(0)
    return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=f"Archive_{wagon_number}_{get_moscow_now().strftime('%Y%m%d_%H%M%S')}.xlsx")


@export_bp.route('/export_active_wagons_excel')
def export_active_wagons_excel():
    station_id = g.get('station_id', 1)
    conn = get_conn()
    
    # Получаем все активные вагоны с их визитами
    df = pd.read_sql_query("""
        SELECT 
            wv.id as visit_id,
            wv.wagon_number as "Номер вагона", 
            wv.owner as "Транспортная компания", 
            wv.organization as "Организация", 
            wv.arrival_time as "Время прибытия", 
            wv.departure_time as "Глобальный срок" 
        FROM wagon_visits wv 
        WHERE wv.status != 'departed' AND wv.is_archived = 0 AND wv.station_id = ?
        ORDER BY wv.wagon_number
    """, conn, params=(station_id,))
    conn.close()
    
    if df.empty:
        flash("Нет активных вагонов для экспорта", 'error')
        return redirect(url_for('history.history_page', station_id=station_id))
    
    # Рассчитываем перепростой на текущую дату для каждого вагона
    overstays = []
    amounts = []
    today_str = get_moscow_now().strftime('%Y-%m-%d')
    
    for _, row in df.iterrows():
        visit_id = row['visit_id']
        overstay = calculate_current_overstay(visit_id)
        overstays.append(overstay if overstay > 0 else '')
        
        if overstay > 0:
            progressive = get_setting('overstay_progressive', '0', station_id) == '1'
            if progressive:
                range1_limit = int(get_setting('overstay_range1_limit', '4', station_id))
                range1_rate = float(get_setting('overstay_range1_rate', '2000', station_id))
                range2_limit = int(get_setting('overstay_range2_limit', '7', station_id))
                range2_rate = float(get_setting('overstay_range2_rate', '2400', station_id))
                range3_rate = float(get_setting('overstay_range3_rate', '3000', station_id))
                
                amount = 0.0
                days1 = min(overstay, range1_limit)
                amount += days1 * range1_rate
                remaining = overstay - days1
                if remaining > 0:
                    days2 = min(remaining, range2_limit - range1_limit)
                    amount += days2 * range2_rate
                    remaining -= days2
                    if remaining > 0:
                        amount += remaining * range3_rate
            else:
                fixed_rate = float(get_setting('overstay_fixed_rate', '2000', station_id))
                amount = overstay * fixed_rate
            amounts.append(amount)
        else:
            amounts.append('')
    
    df['Перепростой на текущую дату, сут'] = overstays
    df['Сумма, руб'] = amounts
    df.drop(columns=['visit_id'], inplace=True)
    
    try:
        df['Время прибытия'] = pd.to_datetime(df['Время прибытия']).dt.strftime('%d-%m-%Y %H:%M:%S')
        df['Глобальный срок'] = pd.to_datetime(df['Глобальный срок']).dt.strftime('%d-%m-%Y %H:%M:%S')
    except:
        pass
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Активные вагоны', index=False)
        apply_excel_styling(writer, 'Активные вагоны')
        
        worksheet = writer.sheets['Активные вагоны']
        cell = worksheet.cell(row=1, column=len(df.columns)+2)
        cell.value = f"Перепростой рассчитан на дату: {today_str}"
        cell.font = Font(italic=True, size=10)
        cell.alignment = Alignment(horizontal='left')
    
    output.seek(0)
    return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=f"ActiveWagons_{get_moscow_now().strftime('%Y%m%d_%H%M%S')}.xlsx")