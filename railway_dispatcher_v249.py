# -*- coding: utf-8 -*-
import webbrowser
import os
import sys
import threading
import time
import shutil
import glob
from flask import Flask, render_template_string, request, redirect, url_for, flash, jsonify, send_file, abort
import sqlite3
from datetime import datetime, timedelta
import pandas as pd
import io
import re
from collections import defaultdict

# ==================== ПУТИ (ВАЖНО ДЛЯ EXE) ====================
def get_base_dir():
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    else:
        return os.path.dirname(os.path.abspath(__file__))

BASE_DIR = get_base_dir()
DB_NAME = os.path.join(BASE_DIR, 'rail_yard.db')
BACKUP_DIR = os.path.join(BASE_DIR, 'backups')
CHANGELOG_PATH = os.path.join(BASE_DIR, 'CHANGELOG.txt')

if not os.path.exists(BACKUP_DIR):
    os.makedirs(BACKUP_DIR)

# ==================== НАСТРОЙКИ РЕЗЕРВНОГО КОПИРОВАНИЯ ====================
BACKUP_HOUR = 3
BACKUP_KEEP_COUNT = 30

# ==================== ГЛОБАЛЬНЫЕ ПЕРЕМЕННЫЕ ====================
SERVER_IP = None
APP_VERSION = "2.4.9"

# ==================== ТРЕЙ ====================
try:
    import pystray
    from PIL import Image, ImageDraw, ImageResampling
    HAS_TRAY = True
except ImportError:
    try:
        import pystray
        from PIL import Image, ImageDraw
        if not hasattr(Image, 'Resampling'):
            Image.Resampling = Image
        HAS_TRAY = True
    except ImportError:
        HAS_TRAY = False
        print("Для работы иконки в трее установите: pip install pystray pillow")

app = Flask(__name__)
app.secret_key = 'rail_app_secret_key_change_me'

def get_conn():
    return sqlite3.connect(DB_NAME, timeout=10, check_same_thread=False)

RETURN_TRACK_NAMES = ["Пост №2", "Ст. Черкасов Камень"]

# ==================== ИНИЦИАЛИЗАЦИЯ БД ====================
def init_db():
    conn = get_conn()
    c = conn.cursor()
    c.execute("PRAGMA journal_mode=WAL")
    
    c.execute('''CREATE TABLE IF NOT EXISTS movement_history (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    wagon_number TEXT, action_type TEXT, from_track TEXT, to_track TEXT, note TEXT, timestamp TEXT
                )''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS archived_history (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    wagon_number TEXT, action_type TEXT, from_track TEXT, to_track TEXT, note TEXT, timestamp TEXT, archived_date TEXT
                )''')

    c.execute('''CREATE TABLE IF NOT EXISTS tracks (id INTEGER PRIMARY KEY, name TEXT UNIQUE, total_length REAL, track_type TEXT DEFAULT 'normal')''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS wagons (
                    id INTEGER PRIMARY KEY AUTOINCREMENT, 
                    wagon_number TEXT UNIQUE, length REAL, cargo_type TEXT, owner TEXT, 
                    organization TEXT, status TEXT DEFAULT 'assigned', track_id INTEGER, 
                    start_pos REAL, arrival_time TEXT, departure_time TEXT, local_departure_time TEXT,
                    visit_count INTEGER DEFAULT 0, is_archived INTEGER DEFAULT 0
                )''')
    
    for col in ['organization', 'local_departure_time', 'owner', 'visit_count', 'is_archived']:
        try: c.execute(f"ALTER TABLE wagons ADD COLUMN {col} TEXT")
        except: pass
    
    c.execute('''CREATE TABLE IF NOT EXISTS action_log (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    timestamp TEXT,
                    username TEXT,
                    ip_address TEXT,
                    action TEXT,
                    wagon_number TEXT,
                    details TEXT,
                    old_value TEXT,
                    new_value TEXT
                )''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS ip_users (
                    ip_address TEXT PRIMARY KEY,
                    username TEXT,
                    note TEXT,
                    is_admin INTEGER DEFAULT 0,
                    role TEXT DEFAULT 'dispatcher',
                    access_allowed INTEGER DEFAULT 0
                )''')
    try:
        c.execute("ALTER TABLE ip_users ADD COLUMN is_admin INTEGER DEFAULT 0")
    except:
        pass
    try:
        c.execute("ALTER TABLE ip_users ADD COLUMN role TEXT DEFAULT 'dispatcher'")
    except:
        pass
    try:
        c.execute("ALTER TABLE ip_users ADD COLUMN access_allowed INTEGER DEFAULT 0")
    except:
        pass
    
    c.execute("UPDATE ip_users SET role='admin', access_allowed=1 WHERE is_admin=1 AND (role='dispatcher' OR role='')")
    c.execute("UPDATE ip_users SET access_allowed=1 WHERE is_admin=1")
    c.execute("UPDATE ip_users SET access_allowed=1 WHERE access_allowed=0 AND role='dispatcher'")
    
    c.execute("UPDATE tracks SET name = 'Резерв' WHERE name = 'Очередь (Буфер)'")
    conn.commit()
    
    c.execute("SELECT count(*) FROM tracks")
    if c.fetchone()[0] == 0:
        data = [
            (1, 'Ст. Черкасов Камень', 1000.0, 'normal'),
            (2, 'Пост №2', 1000.0, 'normal'),
            (3, 'АО "Знамя" (Осмотр)', 1000.0, 'normal'),
            (4, 'АО "Знамя" (Ремонт)', 1000.0, 'normal'),
            (5, 'АО "Знамя" (База - Погрузка)', 1000.0, 'normal'),
            (6, 'АО "Знамя" (Цех ППВВ - Погрузка)', 1000.0, 'normal'),
            (7, 'АО "Знамя" (Отстой)', 1000.0, 'normal'),
            (8, 'Резерв', 2000.0, 'normal')
        ]
        c.executemany("INSERT INTO tracks VALUES (?, ?, ?, ?)", data)
        print("[OK] База данных создана.")
    conn.commit()
    conn.close()

def clean_action_log():
    conn = get_conn()
    c = conn.cursor()
    c.execute("DELETE FROM action_log WHERE timestamp < datetime('now', '-6 months')")
    c.execute("SELECT COUNT(*) FROM action_log")
    count = c.fetchone()[0]
    if count > 20000:
        c.execute("DELETE FROM action_log WHERE id NOT IN (SELECT id FROM action_log ORDER BY id DESC LIMIT 20000)")
    conn.commit()
    conn.close()
    print(f"Журнал действий очищен: осталось записей {min(count, 20000)}")

# ==================== РЕЗЕРВНОЕ КОПИРОВАНИЕ (АВТОМАТИЧЕСКОЕ) ====================
def get_last_auto_backup_time():
    auto_dir = os.path.join(BACKUP_DIR, 'auto')
    if not os.path.exists(auto_dir):
        return None
    backups = glob.glob(os.path.join(auto_dir, 'rail_yard_auto_*.db'))
    if not backups:
        return None
    backups.sort(key=os.path.getmtime, reverse=True)
    last_backup = backups[0]
    return datetime.fromtimestamp(os.path.getmtime(last_backup))

def log_action_no_request(action, wagon_number=None, details=None, old_value=None, new_value=None):
    try:
        conn = get_conn()
        c = conn.cursor()
        c.execute('''INSERT INTO action_log 
            (timestamp, username, ip_address, action, wagon_number, details, old_value, new_value)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)''',
            (datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
             'system', '127.0.0.1', action, wagon_number, details, old_value, new_value))
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"Ошибка записи лога: {e}")

def create_auto_backup():
    try:
        auto_dir = os.path.join(BACKUP_DIR, 'auto')
        if not os.path.exists(auto_dir):
            os.makedirs(auto_dir)
        backup_name = f"rail_yard_auto_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
        backup_path = os.path.join(auto_dir, backup_name)
        shutil.copy2(DB_NAME, backup_path)
        all_backups = sorted(glob.glob(os.path.join(auto_dir, 'rail_yard_auto_*.db')), key=os.path.getmtime)
        while len(all_backups) > BACKUP_KEEP_COUNT:
            os.remove(all_backups.pop(0))
        log_action_no_request('backup_auto', details=f"Автоматическая копия: {backup_path}")
        print(f"📦 Автоматический бэкап создан: {backup_path}")
    except Exception as e:
        print(f"⚠️ Ошибка автоматического бэкапа: {e}")

def schedule_daily_backup():
    def backup_loop():
        while True:
            now = datetime.now()
            next_run = now.replace(hour=BACKUP_HOUR, minute=0, second=0, microsecond=0)
            if now >= next_run:
                next_run += timedelta(days=1)
            wait_seconds = (next_run - now).total_seconds()
            print(f"⏰ Следующий автоматический бэкап запланирован на {next_run.strftime('%Y-%m-%d %H:%M:%S')}")
            time.sleep(wait_seconds)
            create_auto_backup()
    thread = threading.Thread(target=backup_loop, daemon=True)
    thread.start()

# ==================== ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ДЛЯ РОЛЕЙ И ДОСТУПА ====================
def is_return_track(track_name):
    return any(rt in track_name for rt in RETURN_TRACK_NAMES)

def clean_note_for_db(note):
    if not note: return ""
    clean = re.sub('<[^<]+?>', '', str(note))
    clean = clean.replace('\n', ' ').replace('\r', ' ')
    return ' '.join(clean.split()).strip()

def get_user_by_ip(ip):
    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT username, role, access_allowed FROM ip_users WHERE ip_address = ?", (ip,))
    row = c.fetchone()
    conn.close()
    if row:
        return row[0], row[1], bool(row[2])
    return None, None, False

def is_ip_allowed(ip):
    if ip in ('127.0.0.1', '::1'):
        return True
    _, _, allowed = get_user_by_ip(ip)
    return allowed

def get_role_by_ip(ip):
    if ip in ('127.0.0.1', '::1'):
        return 'admin'
    _, role, _ = get_user_by_ip(ip)
    return role if role else 'viewer'

def check_access_for_route(ip, endpoint):
    if not is_ip_allowed(ip):
        return False, None, "Доступ запрещён: ваш IP не внесён в белый список."
    role = get_role_by_ip(ip)
    
    public_endpoints = [
        'index', 'history_page', 'archive_page', 'help_page',
        'api_status', 'get_wagon_info', 'api_dashboard_data', 'static',
        'export_excel', 'export_history_excel', 'export_archive_excel',
        'export_wagon_history', 'export_wagon_archive'
    ]
    dispatcher_endpoints = ['add_wagon', 'move_action', 'depart_action']
    supervisor_endpoints = ['edit_wagon_route', 'edit_history']
    admin_endpoints = [
        'create_backup', 'list_backups', 'download_backup', 'restore_backup',
        'view_logs', 'export_logs_excel', 'manage_ip_users', 'changelog'
    ]
    
    if endpoint in public_endpoints:
        return True, role, None
    elif endpoint in dispatcher_endpoints:
        if role in ('dispatcher', 'admin', 'supervisor'):
            return True, role, None
        else:
            return False, role, "Недостаточно прав. Требуется роль диспетчера или выше."
    elif endpoint in supervisor_endpoints:
        if role in ('supervisor', 'admin'):
            return True, role, None
        else:
            return False, role, "Недостаточно прав. Требуется роль супервизора или администратора."
    elif endpoint in admin_endpoints:
        if role == 'admin':
            return True, role, None
        else:
            return False, role, "Доступ только для администратора."
    else:
        return False, role, "Маршрут не доступен."

@app.before_request
def before_request_check():
    if request.endpoint in ('static',):
        return
    ip = request.remote_addr
    allowed, role, msg = check_access_for_route(ip, request.endpoint)
    if not allowed:
        if request.path.startswith('/api/'):
            return jsonify({"error": msg}), 403
        else:
            return f"<html><body><h1>403 Доступ запрещён</h1><p>{msg}</p><p>Ваш IP: {ip}</p><p><a href='/'>На главную</a></p></body></html>", 403
    request.user_role = role

def get_username_by_ip(ip):
    if ip in ('127.0.0.1', '::1'):
        return "admin"
    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT username FROM ip_users WHERE ip_address = ?", (ip,))
    row = c.fetchone()
    conn.close()
    if row:
        return row[0]
    return ip

def log_action(action, wagon_number=None, details=None, old_value=None, new_value=None):
    try:
        ip = request.remote_addr
        username = get_username_by_ip(ip)
        conn = get_conn()
        c = conn.cursor()
        c.execute('''INSERT INTO action_log 
            (timestamp, username, ip_address, action, wagon_number, details, old_value, new_value)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)''',
            (datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
             username, ip, action, wagon_number, details, old_value, new_value))
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"Ошибка записи лога: {e}")

def get_last_event_datetime(wagon_number):
    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT timestamp FROM movement_history WHERE wagon_number = ? ORDER BY timestamp DESC LIMIT 1", (wagon_number,))
    row = c.fetchone()
    if row:
        conn.close()
        return datetime.strptime(row[0], '%Y-%m-%d %H:%M:%S')
    c.execute("SELECT arrival_time FROM wagons WHERE wagon_number = ? AND is_archived = 0", (wagon_number,))
    row = c.fetchone()
    conn.close()
    if row and row[0]:
        return datetime.strptime(row[0], '%Y-%m-%d %H:%M:%S')
    return None

def log_movement(wagon_number, action_type, from_track_name=None, to_track_name=None, note=None, custom_timestamp=None):
    conn = get_conn()
    c = conn.cursor()
    timestamp = custom_timestamp if custom_timestamp else datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    clean_note = clean_note_for_db(note)
    c.execute("""INSERT INTO movement_history (wagon_number, action_type, from_track, to_track, note, timestamp) VALUES (?, ?, ?, ?, ?, ?)""",
              (wagon_number, action_type, from_track_name, to_track_name, clean_note, timestamp))
    conn.commit()
    conn.close()

def compact_track(track_id):
    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT id, length FROM wagons WHERE track_id = ? AND status != 'departed' AND is_archived = 0 ORDER BY start_pos ASC", (track_id,))
    wagons = c.fetchall()
    current_pos = 0.0
    for wag_id, wag_len in wagons:
        w_len = float(wag_len) if wag_len is not None else 10.0
        c.execute("UPDATE wagons SET start_pos = ? WHERE id = ?", (current_pos, wag_id))
        current_pos += w_len + 50.0
    conn.commit()
    conn.close()

def find_slot_on_track(track_id, wagon_length):
    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT total_length FROM tracks WHERE id = ?", (track_id,))
    res = c.fetchone()
    if not res: 
        conn.close()
        return None, 0.0
    c.execute("SELECT start_pos, length FROM wagons WHERE track_id = ? AND status != 'departed' AND is_archived = 0 ORDER BY start_pos", (track_id,))
    occupied = c.fetchall()
    if not occupied: 
        conn.close()
        return track_id, 0.0
    last_wagon = occupied[-1]
    last_pos = float(last_wagon[0]) if last_wagon[0] is not None else 0.0
    last_len = float(last_wagon[1]) if last_wagon[1] is not None else 10.0
    next_pos = last_pos + last_len + 50 
    conn.close()
    return track_id, next_pos

def move_wagon(wagon_id, new_track_id, local_days=0, local_hours=0, local_mins=0, manual_start_str=None, new_note=None):
    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT wagon_number, owner, organization, departure_time, track_id, cargo_type, visit_count, arrival_time FROM wagons WHERE id = ? AND is_archived = 0", (wagon_id,))
    res = c.fetchone()
    if not res: 
        conn.close()
        return False, "Вагон не найден"
    
    w_num, w_owner, org, global_dep, old_track_id, current_note, current_visits, arrival_time_str = res
    last_event_dt = get_last_event_datetime(w_num)
    
    c.execute("SELECT name FROM tracks WHERE id = ?", (old_track_id,))
    from_track_name = c.fetchone()[0]
    c.execute("SELECT name FROM tracks WHERE id = ?", (new_track_id,))
    to_track_name = c.fetchone()[0]
    
    new_visit_count = int(current_visits) if current_visits is not None else 0
    if not is_return_track(to_track_name): 
        new_visit_count += 1
    
    new_local_dep_time = None
    total_mins = (int(local_days) * 24 * 60) + (int(local_hours) * 60) + int(local_mins)
    
    if manual_start_str and manual_start_str.strip():
        try:
            start_dt = datetime.strptime(manual_start_str.replace('T', ' '), '%Y-%m-%d %H:%M')
            if last_event_dt and start_dt <= last_event_dt:
                conn.close()
                return False, f"Дата начала отсчёта ({start_dt.strftime('%Y-%m-%d %H:%M')}) не может быть раньше или равна предыдущему событию ({last_event_dt.strftime('%Y-%m-%d %H:%M')})"
            if arrival_time_str:
                try:
                    arrival_dt_check = datetime.strptime(arrival_time_str, '%Y-%m-%d %H:%M:%S')
                    if start_dt <= arrival_dt_check:
                        conn.close()
                        return False, f"Дата начала отсчёта не может быть раньше или равна времени прибытия ({arrival_dt_check.strftime('%Y-%m-%d %H:%M')})"
                except:
                    pass
            log_timestamp = manual_start_str.replace('T', ' ') + ":00"
        except ValueError:
            log_timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    else:
        log_timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    if total_mins > 0:
        if manual_start_str and manual_start_str.strip():
            try:
                start_dt = datetime.strptime(manual_start_str.replace('T', ' '), '%Y-%m-%d %H:%M')
                if last_event_dt and start_dt <= last_event_dt:
                    conn.close()
                    return False, f"Дата начала отсчёта ({start_dt.strftime('%Y-%m-%d %H:%M')}) не может быть раньше или равна предыдущему событию ({last_event_dt.strftime('%Y-%m-%d %H:%M')})"
                if arrival_time_str:
                    try:
                        arrival_dt_check = datetime.strptime(arrival_time_str, '%Y-%m-%d %H:%M:%S')
                        if start_dt <= arrival_dt_check:
                            conn.close()
                            return False, f"Дата начала отсчёта не может быть раньше или равна времени прибытия ({arrival_dt_check.strftime('%Y-%m-%d %H:%M')})"
                    except:
                        pass
                new_local_dep_time = (start_dt + timedelta(minutes=total_mins)).strftime('%Y-%m-%d %H:%M:%S')
            except ValueError:
                new_local_dep_time = (datetime.now() + timedelta(minutes=total_mins)).strftime('%Y-%m-%d %H:%M:%S')
        else:
            new_local_dep_time = (datetime.now() + timedelta(minutes=total_mins)).strftime('%Y-%m-%d %H:%M:%S')
    else:
        new_local_dep_time = None
    
    compact_track(new_track_id)
    target_track, new_pos = find_slot_on_track(new_track_id, 10)
    if target_track is None: 
        conn.close()
        return False, "Ошибка пути"
    
    update_note = clean_note_for_db(new_note) if (new_note and new_note.strip()) else clean_note_for_db(current_note)
    
    c.execute("""UPDATE wagons SET track_id = ?, start_pos = ?, local_departure_time = ?, cargo_type = ?, visit_count = ? WHERE id = ?""", 
              (target_track, new_pos, new_local_dep_time, update_note, new_visit_count, wagon_id))
    
    conn.commit()
    conn.close()
    log_movement(w_num, 'moved', from_track_name, to_track_name, f"Примечание: {update_note}" if update_note else "", log_timestamp)
    log_action('move', wagon_number=w_num, details=f"с '{from_track_name}' на '{to_track_name}'")
    return True, "Вагон перемещен!"

def depart_wagon(wagon_id):
    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT wagon_number, track_id FROM wagons WHERE id = ? AND is_archived = 0", (wagon_id,))
    res = c.fetchone()
    if res:
        w_num, track_id = res
        c.execute("SELECT name FROM tracks WHERE id = ?", (track_id,))
        track_name_res = c.fetchone()
        track_name = track_name_res[0] if track_name_res else "Неизвестный путь"
        archived_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        c.execute("""INSERT INTO archived_history (wagon_number, action_type, from_track, to_track, note, timestamp, archived_date) 
                     SELECT wagon_number, action_type, from_track, to_track, note, timestamp, ? FROM movement_history WHERE wagon_number = ?""", 
                  (archived_date, w_num))
        
        c.execute("DELETE FROM movement_history WHERE wagon_number = ?", (w_num,))
        c.execute("UPDATE wagons SET status = 'departed', is_archived = 1 WHERE id = ?", (wagon_id,))
        
        conn.commit()
        conn.close()
        compact_track(track_id)
        
        conn_arch = get_conn()
        c_arch = conn_arch.cursor()
        c_arch.execute("""INSERT INTO archived_history (wagon_number, action_type, from_track, to_track, note, timestamp, archived_date) VALUES (?, ?, ?, ?, ?, ?, ?)""", 
                       (w_num, 'departed', track_name, None, "Убран в архив", archived_date, archived_date))
        conn_arch.commit()
        conn_arch.close()
        log_action('depart', wagon_number=w_num, details=f"Убран в архив с пути {track_name}")
        return True
    return False

# ==================== РАСШИРЕННАЯ ФУНКЦИЯ РЕДАКТИРОВАНИЯ ====================
def parse_flexible_date(date_str):
    if not date_str or not date_str.strip():
        return None
    date_str = date_str.strip()
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%Y-%m-%d'):
        try:
            return datetime.strptime(date_str, fmt)
        except:
            pass
    for fmt in ('%d-%m-%Y %H:%M:%S', '%d-%m-%Y %H:%M', '%d-%m-%Y',
                '%d.%m.%Y %H:%M:%S', '%d.%m.%Y %H:%M', '%d.%m.%Y'):
        try:
            return datetime.strptime(date_str, fmt)
        except:
            pass
    digits = re.sub(r'\D', '', date_str)
    if len(digits) == 12:
        try:
            day = int(digits[0:2]); month = int(digits[2:4]); year = int(digits[4:8])
            hour = int(digits[8:10]); minute = int(digits[10:12])
            return datetime(year, month, day, hour, minute)
        except:
            pass
    if len(digits) == 8:
        try:
            day = int(digits[0:2]); month = int(digits[2:4]); year = int(digits[4:8])
            return datetime(year, month, day)
        except:
            pass
    raise ValueError(f"Не удалось распознать дату: {date_str}")

def edit_wagon(wagon_id, new_owner=None, new_org=None, new_note=None,
               new_arrival_time=None, new_global_deadline=None, new_local_deadline=None):
    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT wagon_number, owner, organization, cargo_type, arrival_time, departure_time, local_departure_time, track_id FROM wagons WHERE id = ? AND is_archived = 0", (wagon_id,))
    row = c.fetchone()
    if not row:
        conn.close()
        return False, "Вагон не найден или находится в архиве"
    w_num, old_owner, old_org, old_note, old_arrival, old_global, old_local, track_id = row
    
    updates = []
    params = []
    changes = []
    
    if new_owner is not None and new_owner != old_owner:
        updates.append("owner = ?")
        params.append(new_owner)
        changes.append(f"ТК: '{old_owner}' → '{new_owner}'")
    if new_org is not None and new_org != old_org:
        updates.append("organization = ?")
        params.append(new_org)
        changes.append(f"Организация: '{old_org}' → '{new_org}'")
    if new_note is not None and new_note != old_note:
        updates.append("cargo_type = ?")
        params.append(clean_note_for_db(new_note))
        changes.append(f"Примечание: '{old_note}' → '{new_note}'")
    
    last_event_dt = get_last_event_datetime(w_num)
    
    if new_arrival_time is not None and new_arrival_time.strip() != "":
        try:
            new_arr_dt = parse_flexible_date(new_arrival_time)
            if new_arr_dt is None:
                raise ValueError("Пустая дата")
            new_arrival_time_str = new_arr_dt.strftime('%Y-%m-%d %H:%M:%S')
        except ValueError as e:
            conn.close()
            return False, f"Ошибка в дате прибытия: {e}"
        if last_event_dt and new_arr_dt > last_event_dt:
            conn.close()
            return False, "Дата прибытия не может быть позже последнего перемещения"
        updates.append("arrival_time = ?")
        params.append(new_arrival_time_str)
        changes.append(f"Время прибытия: '{old_arrival}' → '{new_arrival_time_str}'")
    
    if new_global_deadline is not None and new_global_deadline.strip() != "":
        try:
            new_glob_dt = parse_flexible_date(new_global_deadline)
            if new_glob_dt is None:
                raise ValueError("Пустая дата")
            new_global_str = new_glob_dt.strftime('%Y-%m-%d %H:%M:%S')
        except ValueError as e:
            conn.close()
            return False, f"Ошибка в глобальном сроке: {e}"
        arrival_dt = datetime.strptime(old_arrival, '%Y-%m-%d %H:%M:%S') if old_arrival else None
        if arrival_dt and new_glob_dt < arrival_dt:
            conn.close()
            return False, "Глобальный срок не может быть раньше времени прибытия"
        if last_event_dt and new_glob_dt < last_event_dt:
            conn.close()
            return False, "Глобальный срок не может быть раньше последнего перемещения"
        updates.append("departure_time = ?")
        params.append(new_global_str)
        changes.append(f"Глобальный срок: '{old_global}' → '{new_global_str}'")
    
    if new_local_deadline is not None and new_local_deadline.strip() != "":
        try:
            new_local_dt = parse_flexible_date(new_local_deadline)
            if new_local_dt is None:
                raise ValueError("Пустая дата")
            new_local_str = new_local_dt.strftime('%Y-%m-%d %H:%M:%S')
        except ValueError as e:
            conn.close()
            return False, f"Ошибка в локальном сроке: {e}"
        arrival_dt = datetime.strptime(old_arrival, '%Y-%m-%d %H:%M:%S') if old_arrival else None
        if arrival_dt and new_local_dt < arrival_dt:
            conn.close()
            return False, "Локальный срок не может быть раньше времени прибытия"
        if last_event_dt and new_local_dt < last_event_dt:
            conn.close()
            return False, "Локальный срок не может быть раньше последнего перемещения"
        updates.append("local_departure_time = ?")
        params.append(new_local_str)
        changes.append(f"Локальный срок: '{old_local}' → '{new_local_str}'")
    
    if not updates:
        conn.close()
        return True, "Нет изменений"
    
    query = "UPDATE wagons SET " + ", ".join(updates) + " WHERE id = ?"
    params.append(wagon_id)
    c.execute(query, params)
    conn.commit()
    conn.close()
    
    changes_str = "; ".join(changes)
    log_movement(w_num, 'edit', note=f"Изменения: {changes_str}", custom_timestamp=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    log_action('edit', wagon_number=w_num, details=changes_str, old_value=changes_str, new_value=changes_str)
    return True, "Данные вагона обновлены"

# ==================== API ====================
@app.route('/api/status')
def api_status():
    conn = get_conn()
    c = conn.cursor()
    now = datetime.now()
    c.execute("""SELECT w.id, w.local_departure_time, w.departure_time FROM wagons w WHERE w.status != 'departed' AND w.is_archived = 0""")
    rows = c.fetchall()
    conn.close()
    result = []
    for w_id, local_dep_str, global_dep_str in rows:
        timer_text = ""
        if local_dep_str:
            try:
                local_dt = datetime.strptime(local_dep_str, '%Y-%m-%d %H:%M:%S')
                diff = (local_dt - now).total_seconds()
                if diff <= 0: 
                    timer_text = "ПРОСРОЧЕНО!"
                else:
                    m, s = divmod(int(diff), 60)
                    h, m = divmod(m, 60)
                    d, h = divmod(h, 24)
                    timer_text = f"{d}д {h:02d}:{m:02d}:{s:02d}" if d > 0 else f"{h:02d}:{m:02d}:{s:02d}"
            except: 
                pass
        result.append({"id": w_id, "text": timer_text})
    return jsonify({"timers": result})

@app.route('/api/wagon_info')
def get_wagon_info():
    num = request.args.get('num', '').strip()
    if not num: 
        return jsonify({})
    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT owner, organization FROM wagons WHERE wagon_number = ? LIMIT 1", (num,))
    row = c.fetchone()
    if not row: 
        c.execute("SELECT owner, organization FROM wagons WHERE wagon_number = ? AND is_archived = 1 ORDER BY id DESC LIMIT 1", (num,))
        row = c.fetchone()
    conn.close()
    if row: 
        return jsonify({"owner": row[0] or "", "org": row[1] or ""})
    return jsonify({})

@app.route('/api/dashboard_data')
def api_dashboard_data():
    tracks, move_list = get_dashboard_data()
    result = []
    for track in tracks:
        wagons_data = []
        for w in track['wagons']:
            wagons_data.append({
                'id': w['id'],
                'num': w['num'],
                'owner': w['owner'],
                'org': w['org'],
                'note': w['note'],
                'pos': w['pos'],
                'arrival': w['arrival'],
                'loc_iso': w['loc']['iso'],
                'loc_raw': w['loc']['raw'],
                'loc_overdue': w['loc']['overdue'],
                'glob_iso': w['glob']['iso'],
                'glob_raw': w['glob']['raw'],
                'glob_overdue': w['glob']['overdue'],
                'is_return_track': w['is_return_track'],
                'is_highlighted_return': w['is_highlighted_return'],
                'is_global_overdue': w['is_global_overdue']
            })
        result.append({
            'id': track['id'],
            'name': track['name'],
            'total': track['total'],
            'wagons': wagons_data
        })
    return jsonify({'tracks': result, 'total_wagons': len(move_list)})

# ==================== ИСТОРИЯ И АРХИВ ====================
def get_grouped_history():
    conn = get_conn()
    c = conn.cursor()
    c.execute("""SELECT m.id, m.wagon_number, m.action_type, m.from_track, m.to_track, m.note, m.timestamp, w.owner, w.organization, w.cargo_type 
                 FROM movement_history m 
                 LEFT JOIN wagons w ON m.wagon_number = w.wagon_number 
                 ORDER BY m.wagon_number, m.timestamp ASC""")
    rows = c.fetchall()
    conn.close()
    grouped = defaultdict(list)
    for row in rows:
        hist_id, w_num, action, from_t, to_t, note, ts_str, owner, org, cargo_type = row
        if action == 'added':
            action_label = "<span style='color:#27ae60'>Добавлен</span>"
        elif action == 'moved':
            action_label = "<span style='color:#f39c12'>Перемещен</span>"
        elif action == 'edit':
            action_label = "<span style='color:#8e44ad'>Изменён</span>"
        else:
            action_label = "<span style='color:#e74c3c'>Убыл</span>"
        grouped[w_num].append({
            "id": hist_id,
            "action": action_label,
            "from": from_t or "-",
            "to": to_t or "-",
            "owner": owner or "-",
            "org": org or "-",
            "cargo": cargo_type or "-",
            "note": note or "-",
            "time": ts_str
        })
    def sort_key(k):
        try:
            return (0, int(k))
        except:
            return (1, k)
    sorted_wagons = sorted(grouped.keys(), key=sort_key)
    result = []
    for w_num in sorted_wagons:
        events = grouped[w_num]
        # Добавляем флаг is_last для каждого события
        for idx, ev in enumerate(events):
            ev['is_last'] = (idx == len(events) - 1)
        result.append({
            "num": w_num,
            "last_status": events[-1]['action'],
            "last_time": events[-1]['time'],
            "events": events,
            "count": len(events)
        })
    return result

def get_grouped_archive_history():
    conn = get_conn()
    c = conn.cursor()
    c.execute("""SELECT a.wagon_number, a.action_type, a.from_track, a.to_track, a.note, a.timestamp, w.owner, w.organization, w.cargo_type 
                 FROM archived_history a 
                 LEFT JOIN wagons w ON a.wagon_number = w.wagon_number 
                 ORDER BY a.wagon_number, a.timestamp ASC""")
    rows = c.fetchall()
    conn.close()
    grouped = defaultdict(list)
    for row in rows:
        w_num, action, from_t, to_t, note, ts_str, owner, org, cargo_type = row
        if action == 'added':
            action_label = "<span style='color:#27ae60'>Добавлен</span>"
        elif action == 'moved':
            action_label = "<span style='color:#f39c12'>Перемещен</span>"
        else:
            action_label = "<span style='color:#e74c3c'>Убыл</span>"
        grouped[w_num].append({
            "action": action_label, 
            "from": from_t or "-", 
            "to": to_t or "-", 
            "owner": owner or "-", 
            "org": org or "-", 
            "cargo": cargo_type or "-",
            "note": note or "-",
            "time": ts_str
        })
    def sort_key(k):
        try:
            return (0, int(k))
        except:
            return (1, k)
    sorted_wagons = sorted(grouped.keys(), key=sort_key)
    result = []
    for w_num in sorted_wagons:
        events = grouped[w_num]
        result.append({
            "num": w_num, 
            "last_status": events[-1]['action'], 
            "last_time": events[-1]['time'], 
            "events": events, 
            "count": len(events)
        })
    return result

@app.route('/history')
def history_page(): 
    return render_template_string(HISTORY_SPOILER_TEMPLATE, history_groups=get_grouped_history(), title="История перемещений", session_role=request.user_role)

@app.route('/archive')
def archive_page():
    conn = get_conn()
    c = conn.cursor()
    c.execute("""SELECT wagon_number, owner, organization, departure_time FROM wagons WHERE is_archived = 1 ORDER BY wagon_number ASC""")
    meta_rows = c.fetchall()
    meta_dict = {row[0]: {"owner": row[1] or "-", "org": row[2] or "-", "dep": row[3] or "-"} for row in meta_rows}
    history_data = get_grouped_archive_history()
    full_archive_data = []
    for item in history_data:
        meta = meta_dict.get(item['num'], {"owner": "-", "org": "-", "dep": "-"})
        full_archive_data.append({
            "num": item['num'], 
            "owner": meta['owner'], 
            "org": meta['org'], 
            "dep": meta['dep'], 
            "last_status": item['last_status'], 
            "last_time": item['last_time'], 
            "events": item['events'], 
            "count": item['count']
        })
    conn.close()
    return render_template_string(ARCHIVE_SPOILER_TEMPLATE, archive_groups=full_archive_data)

# ==================== ЭКСПОРТ В EXCEL ====================
@app.route('/export_excel')
def export_excel():
    conn = get_conn()
    df = pd.read_sql_query("""
        SELECT 
            w.wagon_number as "Номер вагона", 
            w.owner as "Транспортная компания", 
            w.organization as "Организация", 
            t.name as "Путь", 
            w.arrival_time as "Время прибытия", 
            w.departure_time as "Глобальный срок" 
        FROM wagons w 
        JOIN tracks t ON w.track_id = t.id 
        WHERE w.status != 'departed' AND w.is_archived = 0
        ORDER BY w.wagon_number
    """, conn)
    conn.close()
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Отчет', index=False)
        worksheet = writer.sheets['Отчет']
        for column in worksheet.columns:
            max_length = 0
            col_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[col_letter].width = adjusted_width
        worksheet.auto_filter.ref = worksheet.dimensions
        from openpyxl.styles import Font, Alignment
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
    output.seek(0)
    return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name=f"Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

@app.route('/export_history_excel')
def export_history_excel():
    conn = get_conn()
    df = pd.read_sql_query("""
        SELECT 
            m.wagon_number as "Номер вагона", 
            m.action_type as "Тип действия", 
            m.from_track as "Откуда", 
            m.to_track as "Куда", 
            w.owner as "Транспортная компания",
            w.organization as "Организация",
            m.note as "Примечание", 
            m.timestamp as "Время" 
        FROM movement_history m
        LEFT JOIN wagons w ON m.wagon_number = w.wagon_number
        ORDER BY m.timestamp DESC
    """, conn)
    conn.close()
    action_map = {'added': 'Добавлен', 'moved': 'Перемещен', 'departed': 'Убыл', 'edit': 'Изменён'}
    df['Тип действия'] = df['Тип действия'].map(action_map).fillna(df['Тип действия'])
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='История', index=False)
        worksheet = writer.sheets['История']
        for column in worksheet.columns:
            max_length = 0
            col_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[col_letter].width = adjusted_width
        worksheet.auto_filter.ref = worksheet.dimensions
        from openpyxl.styles import Font, Alignment
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                if cell.column_letter == 'G':
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
    output.seek(0)
    return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name=f"History_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

@app.route('/export_archive_excel')
def export_archive_excel():
    conn = get_conn()
    df_summary = pd.read_sql_query("""
        SELECT 
            w.wagon_number as "Номер вагона", 
            w.owner as "Транспортная компания", 
            w.organization as "Организация", 
            w.departure_time as "Время убытия"
        FROM wagons w 
        WHERE w.is_archived = 1
        ORDER BY w.wagon_number
    """, conn)
    df_details = pd.read_sql_query("""
        SELECT 
            wagon_number as "Номер вагона", 
            action_type as "Тип действия", 
            from_track as "Откуда", 
            to_track as "Куда", 
            note as "Примечание", 
            timestamp as "Время"
        FROM archived_history
        ORDER BY wagon_number, timestamp ASC
    """, conn)
    conn.close()
    action_map = {'added': 'Добавлен', 'moved': 'Перемещен', 'departed': 'Убыл'}
    df_details['Тип действия'] = df_details['Тип действия'].map(action_map).fillna(df_details['Тип действия'])
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_summary.to_excel(writer, sheet_name='Сводка', index=False)
        worksheet_summary = writer.sheets['Сводка']
        for column in worksheet_summary.columns:
            max_length = 0
            col_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet_summary.column_dimensions[col_letter].width = adjusted_width
        worksheet_summary.auto_filter.ref = worksheet_summary.dimensions
        from openpyxl.styles import Font, Alignment
        for cell in worksheet_summary[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for row in worksheet_summary.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
        df_details.to_excel(writer, sheet_name='Детализация', index=False)
        worksheet_details = writer.sheets['Детализация']
        for column in worksheet_details.columns:
            max_length = 0
            col_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet_details.column_dimensions[col_letter].width = adjusted_width
        worksheet_details.auto_filter.ref = worksheet_details.dimensions
        for cell in worksheet_details[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for row in worksheet_details.iter_rows(min_row=2):
            for cell in row:
                if cell.column_letter == 'E':
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
    output.seek(0)
    return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name=f"Archive_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

@app.route('/export_wagon_history/<wagon_number>')
def export_wagon_history(wagon_number):
    conn = get_conn()
    df = pd.read_sql_query("""
        SELECT 
            action_type as "Тип действия", 
            from_track as "Откуда", 
            to_track as "Куда", 
            note as "Примечание", 
            timestamp as "Время" 
        FROM movement_history 
        WHERE wagon_number = ?
        ORDER BY timestamp ASC
    """, conn, params=(wagon_number,))
    conn.close()
    if df.empty:
        flash(f"Нет данных по вагону {wagon_number}", 'error')
        return redirect(url_for('history_page'))
    action_map = {'added': 'Добавлен', 'moved': 'Перемещен', 'departed': 'Убыл', 'edit': 'Изменён'}
    df['Тип действия'] = df['Тип действия'].map(action_map).fillna(df['Тип действия'])
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=f'История {wagon_number}', index=False)
        worksheet = writer.sheets[f'История {wagon_number}']
        for column in worksheet.columns:
            max_length = 0
            col_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[col_letter].width = adjusted_width
        worksheet.auto_filter.ref = worksheet.dimensions
        from openpyxl.styles import Font, Alignment
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                if cell.column_letter == 'D':
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
    output.seek(0)
    return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name=f"History_{wagon_number}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

@app.route('/export_wagon_archive/<wagon_number>')
def export_wagon_archive(wagon_number):
    conn = get_conn()
    df = pd.read_sql_query("""
        SELECT 
            action_type as "Тип действия", 
            from_track as "Откуда", 
            to_track as "Куда", 
            note as "Примечание", 
            timestamp as "Время" 
        FROM archived_history 
        WHERE wagon_number = ?
        ORDER BY timestamp ASC
    """, conn, params=(wagon_number,))
    conn.close()
    if df.empty:
        flash(f"Нет данных по вагону {wagon_number} в архиве", 'error')
        return redirect(url_for('archive_page'))
    action_map = {'added': 'Добавлен', 'moved': 'Перемещен', 'departed': 'Убыл'}
    df['Тип действия'] = df['Тип действия'].map(action_map).fillna(df['Тип действия'])
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=f'Архив {wagon_number}', index=False)
        worksheet = writer.sheets[f'Архив {wagon_number}']
        for column in worksheet.columns:
            max_length = 0
            col_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[col_letter].width = adjusted_width
        worksheet.auto_filter.ref = worksheet.dimensions
        from openpyxl.styles import Font, Alignment
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                if cell.column_letter == 'D':
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
    output.seek(0)
    return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name=f"Archive_{wagon_number}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

# ==================== ДАННЫЕ ДЛЯ ГЛАВНОЙ СТРАНИЦЫ ====================
def format_date(dt_str):
    if not dt_str: 
        return "-"
    try: 
        return datetime.strptime(str(dt_str)[:16], '%Y-%m-%d %H:%M').strftime('%d.%m.%Y %H:%M')
    except: 
        return str(dt_str)[:19]

def get_dashboard_data():
    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT id, name, total_length, track_type FROM tracks ORDER BY id")
    tracks_raw = c.fetchall()
    c.execute("""SELECT w.id, w.wagon_number, w.length, w.cargo_type, w.owner, w.organization, w.track_id, w.start_pos, w.arrival_time, w.departure_time, w.local_departure_time, t.name, w.visit_count 
                 FROM wagons w 
                 JOIN tracks t ON w.track_id = t.id 
                 WHERE w.status != 'departed' AND w.is_archived = 0 
                 ORDER BY t.id, w.start_pos""")
    all_wagons_raw = c.fetchall()
    conn.close()
    tracks_data = []
    now = datetime.now()
    wagons_by_track = {}
    for w in all_wagons_raw:
        tid = w[6]
        if tid not in wagons_by_track: 
            wagons_by_track[tid] = []
        wagons_by_track[tid].append(w)
    for t_id, t_name, t_len, t_type in tracks_raw:
        try: 
            t_len = float(t_len)
        except: 
            t_len = 1000.0
        track_wagons = wagons_by_track.get(t_id, [])
        processed = []
        is_return_track_flag = is_return_track(t_name)
        for w in track_wagons:
            w_id, w_num, w_len, w_note, w_owner, w_org, w_tid, w_pos, w_arr, w_glob, w_loc, tr_name, w_visits = w
            try: 
                w_visits = int(w_visits) if w_visits is not None else 0
            except: 
                w_visits = 0
            try: 
                w_pos = float(w_pos) if w_pos is not None else 0.0
            except: 
                w_pos = 0.0
            try: 
                w_len = float(w_len) if w_len is not None else 10.0
            except: 
                w_len = 10.0
            loc_text_parts = {"d": 0, "h": 0, "m": 0, "s": 0, "raw": 999999, "iso": "", "overdue": False}
            if w_loc:
                try: 
                    dt = datetime.strptime(str(w_loc)[:19], '%Y-%m-%d %H:%M:%S')
                    loc_iso = dt.strftime('%Y-%m-%dT%H:%M:%S')
                    diff = (dt - now).total_seconds()
                    if diff > 0:
                        m, s = divmod(int(diff), 60)
                        h, m = divmod(m, 60)
                        d, h = divmod(h, 24)
                        loc_text_parts = {
                            "raw": diff,
                            "d": d,
                            "h": h,
                            "m": m,
                            "s": s,
                            "iso": loc_iso,
                            "overdue": False
                        }
                    else:
                        loc_text_parts = {"overdue": True, "raw": diff, "iso": loc_iso, "d": 0, "h": 0, "m": 0, "s": 0}
                except:
                    loc_text_parts = {"raw": 999999, "overdue": False, "iso": "", "d": 0, "h": 0, "m": 0, "s": 0}
            else:
                loc_text_parts = {"raw": 999999, "overdue": False, "iso": "", "d": 0, "h": 0, "m": 0, "s": 0}
            glob_text_parts = {"d": 0, "h": 0, "m": 0, "raw": 0, "iso": "", "overdue": False}
            if w_glob:
                try: 
                    dt = datetime.strptime(str(w_glob)[:19], '%Y-%m-%d %H:%M:%S')
                    glob_iso = dt.strftime('%Y-%m-%dT%H:%M:%S')
                    diff = (dt - now).total_seconds()
                    if diff > 0: 
                        m, s = divmod(int(diff), 60)
                        h, m = divmod(m, 60)
                        d, h = divmod(h, 24)
                        glob_text_parts = {
                            "raw": diff, 
                            "d": d, 
                            "h": h, 
                            "m": m, 
                            "iso": glob_iso,
                            "overdue": False
                        }
                    else: 
                        glob_text_parts = {"raw": diff, "iso": glob_iso, "d": 0, "h": 0, "m": 0, "overdue": True}
                except: 
                    pass
            is_global_overdue = glob_text_parts.get('overdue', False)
            processed.append({
                "id": w_id, 
                "num": w_num, 
                "note": w_note or "-", 
                "owner": w_owner or "Не указана", 
                "org": w_org or "Не указано", 
                "pos": w_pos, 
                "arrival": format_date(w_arr), 
                "loc": loc_text_parts, 
                "glob": glob_text_parts, 
                "is_return_track": is_return_track_flag, 
                "is_highlighted_return": is_return_track_flag and (w_visits > 0),
                "is_global_overdue": is_global_overdue
            })
        tracks_data.append({"id": t_id, "name": t_name, "total": t_len, "type": t_type, "wagons": processed})
    move_list = [{"id": w[0], "text": f"{w[1]} [{w[4] or ''}] ({w[5] or ''}) | {w[11]}", "current_note": w[3] or ""} for w in all_wagons_raw]
    return tracks_data, move_list

# ==================== HTML-ШАБЛОНЫ ====================
HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8">
    <title>ЖД Диспетчерская</title>
    <style>
        * { box-sizing: border-box; }
        body { font-family: 'Segoe UI', Arial, sans-serif; background: #f0f2f5; margin: 0; padding: 20px; }
        .container { max-width: 1400px; margin: 0 auto; background: white; padding: 20px; border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }
        h1 { text-align: center; color: #2c3e50; margin-top: 0; }
        .header-actions { display: flex; justify-content: space-between; align-items: center; margin-bottom: 20px; flex-wrap: wrap; gap: 10px; }
        .nav-buttons { display: flex; gap: 10px; flex-wrap: wrap; }
        .btn { color: white; padding: 8px 15px; border: none; border-radius: 5px; text-decoration: none; font-weight: bold; display: inline-flex; align-items: center; gap: 5px; cursor: pointer; }
        .btn-history { background: #2980b9; }
        .btn-excel { background: #217346; }
        .btn-archive { background: #7f8c8d; }
        .btn-help { background: #16a085; }
        .btn-changelog { background: #2c3e50; }
        .btn-backup { background: #27ae60; }
        .btn-add { background: #27ae60; }
        .btn-move { background: #f39c12; }
        .controls { display: flex; gap: 20px; margin-bottom: 30px; background: #ecf0f1; padding: 20px; border-radius: 8px; flex-wrap: wrap; }
        .control-box { flex: 1; min-width: 300px; }
        .control-box:first-child { border-right: 1px solid #bdc3c7; padding-right: 20px; }
        .control-box:last-child { padding-left: 20px; }
        input, select, textarea { padding: 10px; margin: 5px 0; border: 1px solid #ddd; border-radius: 4px; width: 100%; }
        .time-group { display: flex; gap: 10px; align-items: flex-end; }
        .time-group div { flex: 1; text-align: center; }
        .time-group input { text-align: center; margin-bottom: 3px; }
        .time-group label { font-size: 11px; color: #555; display: block; margin-top: 2px; }
        .manual-time-inputs { display: flex; gap: 5px; align-items: center; flex-wrap: wrap; }
        .manual-time-inputs input { width: 120px; }
        .date-help { font-size: 10px; color: #7f8c8d; margin-top: 2px; }
        .date-buttons { display: flex; gap: 5px; margin-left: 10px; }
        .date-buttons button { background: #ecf0f1; border: 1px solid #bdc3c7; border-radius: 3px; padding: 4px 8px; cursor: pointer; font-size: 11px; }
        .alert { padding: 10px; margin-bottom: 15px; border-radius: 4px; text-align: center; font-weight: bold; }
        .alert-success { background: #d4edda; color: #155724; border-left: 4px solid #27ae60; }
        .alert-error { background: #f8d7da; color: #721c24; border-left: 4px solid #e74c3c; }
        .track-wrapper { margin-bottom: 20px; border: 1px solid #ddd; border-radius: 8px; background: white; }
        .track-header { background: #34495e; color: white; padding: 10px 15px; font-weight: bold; display: flex; justify-content: space-between; border-radius: 8px 8px 0 0; }
        .track-body { position: relative; height: 200px; background: #ecf0f1; margin: 15px; border-bottom: 3px solid #7f8c8d; overflow-x: auto; }
        .wagon { position: absolute; top: 20px; height: 160px; border: 2px solid #2c3e50; border-radius: 6px; color: white; display: flex; flex-direction: column; align-items: center; justify-content: center; font-size: 13px; font-weight: bold; cursor: pointer; box-shadow: 2px 2px 5px rgba(0,0,0,0.2); min-width: 130px; padding: 8px; }
        .wagon:hover { transform: scale(1.02); z-index: 100; }
        .wagon.active { outline: 3px solid #f1c40f; }
        .wagon-normal { background: linear-gradient(135deg, #3498db, #2980b9); }
        .wagon-return-highlight { background: linear-gradient(135deg, #8e44ad, #9b59b6); border: 2px solid #f1c40f; }
        .wagon-global-overdue { background: linear-gradient(135deg, #6c3483, #4a235a); border: 2px solid #e74c3c; }
        .wagon-global-overdue-normal { background: linear-gradient(135deg, #922b21, #641e16); border: 2px solid #e74c3c; }
        .wagon-overdue { background: linear-gradient(135deg, #e74c3c, #c0392b); }
        .wagon-warn { background: linear-gradient(135deg, #f39c12, #d35400); }
        .wagon-number { font-size: 18px; font-weight: bold; text-align: center; margin-bottom: 8px; }
        .wagon-timer { font-family: monospace; font-size: 13px; margin-top: 5px; text-align: center; font-weight: bold; }
        .timer-label { font-size: 11px; opacity: 0.9; margin-right: 4px; }
        .total-count { background: #34495e; color: white; padding: 8px 15px; border-radius: 20px; font-weight: bold; }
        .refresh-btn { background: #95a5a6; border: none; padding: 5px 10px; border-radius: 4px; color: white; cursor: pointer; }
        #global-tooltip { display: none; position: fixed; background: #2c3e50; color: white; padding: 15px; border-radius: 8px; z-index: 1000; min-width: 280px; box-shadow: 0 5px 20px rgba(0,0,0,0.3); }
        .tooltip-row { display: flex; justify-content: space-between; margin-bottom: 8px; padding-bottom: 5px; border-bottom: 1px solid #465c71; }
        .timer-block { margin-top: 10px; padding-top: 10px; border-top: 1px dashed #7f8c8d; }
        .timer-title { font-size: 12px; text-transform: uppercase; color: #bdc3c7; margin-bottom: 5px; }
        .btn-remove-large { background: #e74c3c; color: white; border: none; padding: 8px; border-radius: 4px; cursor: pointer; width: 100%; margin-top: 10px; font-weight: bold; }
        .close-tooltip { text-align: center; margin-top: 8px; font-size: 12px; color: #bdc3c7; cursor: pointer; }
        @media (max-width: 768px) { .control-box:first-child { border-right: none; padding-right: 0; } .control-box:last-child { padding-left: 0; } }
        #moveWagonSelect { height: auto; min-height: 120px; }
        #wagonSearchInput { margin-bottom: 5px; }
        .legend { background: #f8f9fa; padding: 10px; border-radius: 8px; margin-bottom: 20px; display: flex; flex-wrap: wrap; gap: 15px; align-items: center; font-size: 13px; }
        .legend-color { display: inline-block; width: 20px; height: 20px; border-radius: 4px; vertical-align: middle; margin-right: 4px; }
        .date-time-group { display: flex; flex-wrap: wrap; align-items: center; gap: 5px; margin-bottom: 5px; }
        .date-input { font-family: monospace; }
    </style>
</head>
<body>
<div class="container">
<div class="header-actions" style="display: flex; justify-content: space-between; align-items: center; flex-wrap: wrap;">
    <div>
        <h1 style="margin: 0;">🚂 ЖД Диспетчерская</h1>
        <div style="font-size: 11px; color: #7f8c8d; margin-top: 2px;">Версия {{ version }}</div>
    </div>
    <div class="nav-buttons" style="display: flex; gap: 10px; align-items: center;">
        <div class="total-count">📦 Всего: {{ total_wagons }}</div>
        <a href="/history" class="btn btn-history">📜 История</a>
        <a href="/archive" class="btn btn-archive">🗄️ Архив</a>
        <a href="/export_excel" class="btn btn-excel">📊 Excel</a>
        {% if is_admin %}
            <a href="/changelog" class="btn btn-changelog">📋 Список изменений</a>
            <a href="/admin/backups" class="btn btn-backup">💾 Бэкапы</a>
            <a href="/admin/logs" class="btn" style="background:#8e44ad;">📜 Журнал действий</a>
            <a href="/admin/ip_users" class="btn" style="background:#e67e22;">🔗 Привязка IP</a>
        {% endif %}
        <a href="/help" class="btn btn-help">❓ Помощь</a>
    </div>
</div>
    {% with messages = get_flashed_messages(with_categories=true) %}
        {% if messages %}
            {% for cat, msg in messages %}
                <div class="alert alert-{{ cat }}">{% if cat == 'success' %}✅{% else %}⚠️{% endif %} {{ msg }}</div>
            {% endfor %}
        {% endif %}
    {% endwith %}
    <div class="legend">
        <div><span class="legend-color" style="background: linear-gradient(135deg,#3498db,#2980b9);"></span> Синий – обычный вагон (срок в норме или не задан)</div>
        <div><span class="legend-color" style="background: linear-gradient(135deg,#8e44ad,#9b59b6); border:1px solid #f1c40f;"></span> Фиолетовый – был на возвратном пути</div>
        <div><span class="legend-color" style="background: linear-gradient(135deg,#6c3483,#4a235a); border:2px solid #e74c3c;"></span> Тёмно-фиолетовый – был на возвратном пути И истёк глобальный срок</div>
        <div><span class="legend-color" style="background: linear-gradient(135deg,#922b21,#641e16); border:2px solid #e74c3c;"></span> Тёмно-красный – обычный вагон с истёкшим глобальным сроком</div>
        <div><span class="legend-color" style="background: linear-gradient(135deg,#f39c12,#d35400);"></span> Оранжевый – локальный срок &lt; 1 часа</div>
        <div><span class="legend-color" style="background: linear-gradient(135deg,#e74c3c,#c0392b);"></span> Красный – локальный срок истёк (просрочка)</div>
    </div>
    <div class="controls">
        <div class="control-box">
            <h3>➕ Новый вагон</h3>
            <form action="/add" method="POST" onsubmit="return validateDateTime('add')">
                <input type="text" id="new-wagon-number" name="number" placeholder="№ Вагона" required onblur="fetchWagonInfo(this.value)" value="{{ add_form_data.number if add_form_data else '' }}">
                <input type="text" id="new-wagon-owner" name="owner" placeholder="Транспортная компания" required value="{{ add_form_data.owner if add_form_data else '' }}">
                <input type="text" id="new-wagon-org" name="organization" placeholder="Организация" required value="{{ add_form_data.org if add_form_data else '' }}">
                <textarea name="note" rows="2" placeholder="Примечание">{{ add_form_data.note if add_form_data else '' }}</textarea>
                <div style="background:#e8f6f3;padding:10px;border-radius:5px;margin-top:10px">
                    <label style="font-weight:bold;">⏰ Глобальный срок:</label>
                    <div class="time-group">
                        <div><input type="number" name="cycle_days" min="0" value="{{ add_form_data.cycle_days if add_form_data else '0' }}"><label>Дни</label></div>
                        <div><input type="number" name="cycle_hours" min="0" value="{{ add_form_data.cycle_hours if add_form_data else '0' }}"><label>Часы</label></div>
                        <div><input type="number" name="cycle_mins" min="0" value="{{ add_form_data.cycle_mins if add_form_data else '0' }}"><label>Минуты</label></div>
                    </div>
                    <div style="margin-top:10px">
                        <label>📅 Начало отсчета (оставьте пустым для "сейчас"):</label>
                        <div class="date-time-group">
                            <input type="text" id="add_start_date" name="start_date" placeholder="ДД.ММ.ГГГГ" class="date-input" onblur="formatDateInput(this)" autocomplete="off" value="{{ add_form_data.start_date if add_form_data else '' }}">
                            <input type="text" id="add_start_time" name="start_time" placeholder="ЧЧ:ММ или ЧЧММ" class="date-input" onblur="formatTimeInput(this)" autocomplete="off" value="{{ add_form_data.start_time if add_form_data else '' }}">
                            <div class="date-buttons">
                                <button type="button" onclick="setToday('add')">Сегодня</button>
                                <button type="button" onclick="setNow('add')">Сейчас</button>
                                <button type="button" onclick="clearDateTime('add')">Очистить</button>
                            </div>
                        </div>
                        <div class="date-help">Формат: ДД.ММ.ГГГГ (можно без точек) и ЧЧ:ММ (можно без двоеточия). Оба поля либо заполнены, либо пусты.</div>
                    </div>
                </div>
                <select name="track_id" required>
                    <option value="">Выберите путь...</option>
                    {% for t in tracks %}
                        <option value="{{ t.id }}" {% if add_form_data and add_form_data.track_id|string == t.id|string %}selected{% endif %}>{{ t.name }}</option>
                    {% endfor %}
                </select>
                <button type="submit" class="btn-add">➕ Распределить</button>
            </form>
        </div>
        <div class="control-box">
            <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:10px">
                <h3 style="margin:0">🔄 Переместить</h3>
                <button type="button" onclick="location.reload()" class="refresh-btn">🔄 Обновить</button>
            </div>
            <form action="/move" method="POST" onsubmit="return validateDateTime('move')">
                <input type="text" id="wagonSearchInput" placeholder="🔍 Поиск вагона..." onkeyup="filterWagons()">
                <select name="wagon_id" id="moveWagonSelect" required onchange="updateMoveNote()" size="5">
                    <option value="">Выберите вагон...</option>
                    {% for w in move_list %}
                        <option value="{{ w.id }}" data-note="{{ w.current_note }}" {% if move_form_data and move_form_data.wagon_id|string == w.id|string %}selected{% endif %}>{{ w.text }}</option>
                    {% endfor %}
                </select>
                <textarea name="note" id="moveNoteArea" rows="2" placeholder="Примечание...">{{ move_form_data.note if move_form_data else '' }}</textarea>
                <div style="background:#fef9e7;padding:10px;border-radius:5px;margin-top:10px">
                    <label style="font-weight:bold;">⏳ Локальный срок:</label>
                    <div class="time-group">
                        <div><input type="number" name="local_days" min="0" value="{{ move_form_data.local_days if move_form_data else '0' }}"><label>Дни</label></div>
                        <div><input type="number" name="local_hours" min="0" value="{{ move_form_data.local_hours if move_form_data else '0' }}"><label>Часы</label></div>
                        <div><input type="number" name="local_mins" min="0" value="{{ move_form_data.local_mins if move_form_data else '0' }}"><label>Минуты</label></div>
                    </div>
                    <div style="margin-top:10px">
                        <label>📅 Начало отсчета (оставьте пустым для "сейчас"):</label>
                        <div class="date-time-group">
                            <input type="text" id="move_start_date" name="start_date" placeholder="ДД.ММ.ГГГГ" class="date-input" onblur="formatDateInput(this)" autocomplete="off" value="{{ move_form_data.start_date if move_form_data else '' }}">
                            <input type="text" id="move_start_time" name="start_time" placeholder="ЧЧ:ММ или ЧЧММ" class="date-input" onblur="formatTimeInput(this)" autocomplete="off" value="{{ move_form_data.start_time if move_form_data else '' }}">
                            <div class="date-buttons">
                                <button type="button" onclick="setToday('move')">Сегодня</button>
                                <button type="button" onclick="setNow('move')">Сейчас</button>
                                <button type="button" onclick="clearDateTime('move')">Очистить</button>
                            </div>
                        </div>
                        <div class="date-help">Формат: ДД.ММ.ГГГГ (можно без точек) и ЧЧ:ММ (можно без двоеточия). Оба поля либо заполнены, либо пусты.</div>
                    </div>
                </div>
                <select name="new_track_id" required>
                    <option value="">На путь...</option>
                    {% for t in tracks %}
                        <option value="{{ t.id }}" {% if move_form_data and move_form_data.new_track_id|string == t.id|string %}selected{% endif %}>{{ t.name }}</option>
                    {% endfor %}
                </select>
                <button type="submit" class="btn-move">🔄 Переместить</button>
            </form>
        </div>
    </div>
    <div id="global-tooltip">
        <div class="tooltip-row"><span># Номер:</span><span id="tt-num"></span></div>
        <div class="tooltip-row"><span>🚚 ТК:</span><span id="tt-owner"></span></div>
        <div class="tooltip-row"><span>🏢 Орг:</span><span id="tt-org"></span></div>
        <div class="tooltip-row"><span>📝 Примечание:</span><span id="tt-note"></span></div>
        <div class="tooltip-row"><span>📅 Прибыл:</span><span id="tt-arr"></span></div>
        <div class="timer-block"><div class="timer-title">⏰ Локальный срок:</div><div id="tt-loc"></div></div>
        <div class="timer-block"><div class="timer-title">🌍 Глобальный срок:</div><div id="tt-glob"></div></div>
        <div id="tt-remove-container" style="margin-top:15px;text-align:center">
            <form id="tt-depart-form" method="POST" onsubmit="return confirm('Убрать в архив?')">
                <button type="submit" class="btn-remove-large">🗄️ УБРАТЬ В АРХИВ</button>
            </form>
        </div>
        <div id="tt-edit-container" style="margin-top:10px;text-align:center; border-top:1px solid #465c71; padding-top:10px;">
            <button type="button" class="btn-edit" style="background:#f39c12; color:white; border:none; padding:8px; border-radius:4px; cursor:pointer; width:100%;" onclick="openEditModal()">✏️ Редактировать вагон</button>
        </div>
        <div class="close-tooltip" onclick="closeTooltip()">❌ Закрыть</div>
    </div>
    {% for track in tracks %}
        <div class="track-wrapper" data-track-id="{{ track.id }}">
            <div class="track-header"><span>🛤️ {{ track.name }}</span><span>📦 Вагонов: {{ track.wagons|length }}</span></div>
            <div class="track-body">
                {% for w in track.wagons %}
                    {% set left_pct = (w.pos / track.total) * 100 %}
                    {% set is_loc_overdue = w.loc.overdue if w.loc.overdue is defined else false %}
                    {% set raw_loc_time = w.loc.raw if w.loc.raw is defined and w.loc.raw is number else 999999 %}
                    {% set is_global_overdue = w.is_global_overdue if w.is_global_overdue is defined else false %}
                    {% if is_loc_overdue %}
                        {% set wagon_class = "wagon wagon-overdue" %}
                    {% elif raw_loc_time < 3600 %}
                        {% set wagon_class = "wagon wagon-warn" %}
                    {% elif w.is_highlighted_return and is_global_overdue %}
                        {% set wagon_class = "wagon wagon-global-overdue" %}
                    {% elif w.is_highlighted_return %}
                        {% set wagon_class = "wagon wagon-return-highlight" %}
                    {% elif is_global_overdue %}
                        {% set wagon_class = "wagon wagon-global-overdue-normal" %}
                    {% else %}
                        {% set wagon_class = "wagon wagon-normal" %}
                    {% endif %}
                    <div class="{{ wagon_class }}" id="wagon-{{ w.id }}"
                         style="width:130px; left: {{ left_pct }}%;"
                         data-id="{{ w.id }}"
                         data-num="{{ w.num }}"
                         data-owner="{{ w.owner }}"
                         data-org="{{ w.org }}"
                         data-note="{{ w.note }}"
                         data-arrival="{{ w.arrival }}"
                         data-loc-iso="{{ w.loc.iso }}"
                         data-glob-iso="{{ w.glob.iso }}"
                         data-track-name="{{ track.name }}"
                         data-loc-raw="{{ w.loc.raw }}"
                         data-glob-raw="{{ w.glob.raw }}">
                        <div class="wagon-number">{{ w.num }}</div>
                        {% if w.loc.iso %}
                            <div class="wagon-timer" id="timer-loc-{{ w.id }}" data-time="{{ w.loc.iso }}" data-raw="{{ w.loc.raw }}">
                                <span class="timer-label">⏳</span><span>{{ w.loc.d }}д {{ "%02d"|format(w.loc.h) }}:{{ "%02d"|format(w.loc.m) }}:{{ "%02d"|format(w.loc.s) }}</span>
                            </div>
                        {% endif %}
                        {% if w.glob.iso %}
                            <div class="wagon-timer" id="timer-glob-{{ w.id }}" data-time="{{ w.glob.iso }}" data-raw="{{ w.glob.raw }}">
                                <span class="timer-label">🌍</span><span>{{ w.glob.d }}д {{ "%02d"|format(w.glob.h) }}:{{ "%02d"|format(w.glob.m) }}</span>
                            </div>
                        {% endif %}
                    </div>
                {% endfor %}
            </div>
        </div>
    {% endfor %}
</div>
<div id="editModal" style="display:none; position:fixed; top:0; left:0; width:100%; height:100%; background:rgba(0,0,0,0.5); z-index:2000; justify-content:center; align-items:center;">
    <div style="background:white; padding:20px; border-radius:10px; width:500px; max-width:90%;">
        <h3>✏️ Редактирование вагона</h3>
        <form id="editForm">
            <input type="hidden" id="edit_wagon_id">
            <label>Транспортная компания:</label>
            <input type="text" id="edit_owner" name="owner">
            <label>Организация:</label>
            <input type="text" id="edit_org" name="organization">
            <label>Примечание:</label>
            <textarea id="edit_note" name="note" rows="2"></textarea>
            <label>Время прибытия (ГГГГ-ММ-ДД ЧЧ:ММ:СС):</label>
            <input type="text" id="edit_arrival" name="arrival_time" placeholder="2026-04-10 14:30:00">
            <label>Глобальный срок (ГГГГ-ММ-ДД ЧЧ:ММ:СС):</label>
            <input type="text" id="edit_global" name="departure_time" placeholder="2026-04-12 14:30:00">
            <label>Локальный срок (ГГГГ-ММ-ДД ЧЧ:ММ:СС):</label>
            <input type="text" id="edit_local" name="local_departure_time" placeholder="2026-04-11 14:30:00">
            <div style="margin-top:15px; text-align:right;">
                <button type="button" onclick="closeEditModal()" style="padding:5px 10px;">Отмена</button>
                <button type="submit" style="padding:5px 10px; background:#27ae60; color:white; border:none;">Сохранить</button>
            </div>
        </form>
    </div>
</div>
<script>
let activeWagonId = null;
let activeWagonNum = null;
let activeWagonOwner = null;
let activeWagonOrg = null;
let activeWagonNote = null;
let activeWagonArrival = null;
let activeWagonGlobal = null;
let activeWagonLocal = null;
const tooltip = document.getElementById('global-tooltip');
const removeContainer = document.getElementById('tt-remove-container');
const editContainer = document.getElementById('tt-edit-container');

function normalizeDateTimeForEdit(dateTimeStr) {
    if (!dateTimeStr || dateTimeStr.trim() === '') return '';
    let str = dateTimeStr.trim();
    if (/^\\d{4}-\\d{2}-\\d{2} \\d{2}:\\d{2}:\\d{2}$/.test(str)) return str;
    if (/^\\d{4}-\\d{2}-\\d{2}$/.test(str)) return str + ' 00:00:00';
    let match = str.match(/^(\\d{2})[.\\-](\\d{2})[.\\-](\\d{4})(?: (\\d{2}):(\\d{2})(?::(\\d{2}))?)?$/);
    if (match) {
        let day = match[1], month = match[2], year = match[3];
        let hour = match[4] || '00', minute = match[5] || '00', second = match[6] || '00';
        return `${year}-${month}-${day} ${hour}:${minute}:${second}`;
    }
    let digits = str.replace(/[^\\d]/g, '');
    if (digits.length === 8) {
        let day = digits.slice(0,2), month = digits.slice(2,4), year = digits.slice(4,8);
        return `${year}-${month}-${day} 00:00:00`;
    }
    if (digits.length === 12) {
        let day = digits.slice(0,2), month = digits.slice(2,4), year = digits.slice(4,8);
        let hour = digits.slice(8,10), minute = digits.slice(10,12);
        return `${year}-${month}-${day} ${hour}:${minute}:00`;
    }
    return str;
}

function formatDateInput(input) {
    let value = input.value.trim();
    if (value === '') return;
    let digits = value.replace(/[^\\d]/g, '');
    if (digits.length === 8) {
        let day = digits.substring(0, 2);
        let month = digits.substring(2, 4);
        let year = digits.substring(4, 8);
        input.value = `${year}-${month}-${day}`;
    } else if (digits.length === 6) {
        let day = digits.substring(0, 2);
        let month = digits.substring(2, 4);
        let year = digits.substring(4, 6);
        input.value = `20${year}-${month}-${day}`;
    } else if (value.includes('.') && value.split('.').length === 3) {
        let parts = value.split('.');
        if (parts[0].length === 2 && parts[1].length === 2 && parts[2].length === 4) {
            input.value = `${parts[2]}-${parts[1]}-${parts[0]}`;
        }
    } else if (value.includes('-') && value.split('-').length === 3) {
        return;
    }
}

function formatTimeInput(input) {
    let value = input.value.trim();
    if (value === '') return;
    let digits = value.replace(/[^\\d]/g, '');
    if (digits.length === 4) {
        let hours = digits.substring(0, 2);
        let minutes = digits.substring(2, 4);
        if (parseInt(hours) <= 23 && parseInt(minutes) <= 59) {
            input.value = `${hours}:${minutes}`;
        }
    } else if (digits.length === 3) {
        let hours = digits.substring(0, 1);
        let minutes = digits.substring(1, 3);
        if (parseInt(hours) <= 23 && parseInt(minutes) <= 59) {
            input.value = `0${hours}:${minutes}`;
        }
    } else if (digits.length === 2) {
        let hours = digits;
        if (parseInt(hours) <= 23) {
            input.value = `${hours}:00`;
        }
    } else if (digits.length === 1) {
        let hours = digits;
        if (parseInt(hours) <= 23) {
            input.value = `0${hours}:00`;
        }
    } else if (value.includes(':') && value.split(':').length === 2) {
        return;
    }
}

function setToday(formType) {
    const today = new Date();
    const year = today.getFullYear();
    const month = String(today.getMonth() + 1).padStart(2, '0');
    const day = String(today.getDate()).padStart(2, '0');
    const dateValue = `${year}-${month}-${day}`;
    if (formType === 'add') {
        document.getElementById('add_start_date').value = dateValue;
    } else {
        document.getElementById('move_start_date').value = dateValue;
    }
}

function setNow(formType) {
    const now = new Date();
    const hours = String(now.getHours()).padStart(2, '0');
    const minutes = String(now.getMinutes()).padStart(2, '0');
    const timeValue = `${hours}:${minutes}`;
    if (formType === 'add') {
        document.getElementById('add_start_time').value = timeValue;
    } else {
        document.getElementById('move_start_time').value = timeValue;
    }
}

function clearDateTime(formType) {
    if (formType === 'add') {
        document.getElementById('add_start_date').value = '';
        document.getElementById('add_start_time').value = '';
    } else {
        document.getElementById('move_start_date').value = '';
        document.getElementById('move_start_time').value = '';
    }
}

function validateDateTime(formType) {
    let dateField, timeField;
    if (formType === 'add') {
        dateField = document.getElementById('add_start_date');
        timeField = document.getElementById('add_start_time');
    } else {
        dateField = document.getElementById('move_start_date');
        timeField = document.getElementById('move_start_time');
    }
    const hasDate = dateField && dateField.value.trim() !== '';
    const hasTime = timeField && timeField.value.trim() !== '';
    if ((hasDate && !hasTime) || (!hasDate && hasTime)) {
        alert('Ошибка: Если вы заполняете дату, нужно заполнить и время, и наоборот. Либо оставьте оба поля пустыми (будет использовано текущее время).');
        return false;
    }
    if (hasDate && hasTime) {
        const datePattern = /^\\d{4}-\\d{2}-\\d{2}$/;
        const timePattern = /^\\d{2}:\\d{2}$/;
        if (!datePattern.test(dateField.value.trim())) {
            alert('Неверный формат даты. Используйте ДД.ММ.ГГГГ (или просто цифрами).');
            return false;
        }
        if (!timePattern.test(timeField.value.trim())) {
            alert('Неверный формат времени. Используйте ЧЧ:ММ (или просто ЧЧММ).');
            return false;
        }
    }
    return true;
}

function filterWagons() {
    const input = document.getElementById('wagonSearchInput');
    const filter = input.value.toLowerCase();
    const select = document.getElementById('moveWagonSelect');
    const options = select.options;
    let hasVisible = false;
    for (let i = 0; i < options.length; i++) {
        const text = options[i].text.toLowerCase();
        if (text.includes(filter) || filter === '') {
            options[i].style.display = '';
            hasVisible = true;
        } else {
            options[i].style.display = 'none';
        }
    }
    let noResultMsg = document.getElementById('noResultMsg');
    if (!hasVisible && filter !== '') {
        if (!noResultMsg) {
            const msg = document.createElement('div');
            msg.id = 'noResultMsg';
            msg.style.color = '#e74c3c';
            msg.style.fontSize = '12px';
            msg.style.marginTop = '5px';
            msg.innerText = '❌ Ничего не найдено';
            select.parentNode.insertBefore(msg, select.nextSibling);
        }
    } else if (noResultMsg) {
        noResultMsg.remove();
    }
}

function updateMoveNote() {
    const select = document.getElementById('moveWagonSelect');
    const selectedOption = select.options[select.selectedIndex];
    const note = selectedOption ? selectedOption.getAttribute('data-note') : '';
    document.getElementById('moveNoteArea').value = note || '';
}

function openTooltip(el, id, num, owner, org, note, arrival, locIso, globIso, trackName, locRaw, globRaw) {
    if (activeWagonId === id) { closeTooltip(); return; }
    document.getElementById('tt-num').innerText = num;
    document.getElementById('tt-owner').innerText = owner || '-';
    document.getElementById('tt-org').innerText = org || '-';
    document.getElementById('tt-note').innerHTML = note || '-';
    document.getElementById('tt-arr').innerText = arrival;
    const ttLoc = document.getElementById('tt-loc');
    const ttGlob = document.getElementById('tt-glob');
    if (locIso) {
        ttLoc.setAttribute('data-time', locIso);
        ttLoc.setAttribute('data-raw', locRaw);
        ttLoc.innerText = '...';
    } else {
        ttLoc.removeAttribute('data-time');
        ttLoc.innerText = 'Нет';
    }
    if (globIso) {
        ttGlob.setAttribute('data-time', globIso);
        ttGlob.setAttribute('data-raw', globRaw);
        ttGlob.innerText = '...';
    } else {
        ttGlob.removeAttribute('data-time');
        ttGlob.innerText = 'Нет';
    }
    document.getElementById('tt-depart-form').action = '/depart/' + id;
    activeWagonId = id;
    activeWagonNum = num;
    activeWagonOwner = owner;
    activeWagonOrg = org;
    activeWagonNote = note;
    activeWagonArrival = arrival;
    activeWagonGlobal = globIso ? globIso.replace('T', ' ') : '';
    activeWagonLocal = locIso ? locIso.replace('T', ' ') : '';
    document.querySelectorAll('.wagon').forEach(w => w.classList.remove('active'));
    el.classList.add('active');
    const allowedTracks = ["Ст. Черкасов Камень", "Пост №2"];
    if (allowedTracks.includes(trackName.trim())) {
        removeContainer.style.display = 'block';
    } else {
        removeContainer.style.display = 'none';
    }
    editContainer.style.display = 'block';
    const rect = el.getBoundingClientRect();
    let left = rect.right + 10;
    let top = rect.top;
    if (left + 300 > window.innerWidth) left = rect.left - 310;
    if (top + 400 > window.innerHeight) top = window.innerHeight - 410;
    tooltip.style.left = left + 'px';
    tooltip.style.top = top + 'px';
    tooltip.style.display = 'block';
    updateTooltipTimers();
}

function closeTooltip() {
    tooltip.style.display = 'none';
    if (activeWagonId) {
        const w = document.getElementById('wagon-' + activeWagonId);
        if (w) w.classList.remove('active');
    }
    activeWagonId = null;
}

function openEditModal() {
    if (!activeWagonId) return;
    document.getElementById('edit_wagon_id').value = activeWagonId;
    document.getElementById('edit_owner').value = activeWagonOwner || '';
    document.getElementById('edit_org').value = activeWagonOrg || '';
    document.getElementById('edit_note').value = activeWagonNote === '-' ? '' : activeWagonNote;
    document.getElementById('edit_arrival').value = activeWagonArrival && activeWagonArrival !== '-' ? activeWagonArrival.replace(/\\./g, '-') : '';
    document.getElementById('edit_global').value = activeWagonGlobal || '';
    document.getElementById('edit_local').value = activeWagonLocal || '';
    document.getElementById('editModal').style.display = 'flex';
}

function closeEditModal() {
    document.getElementById('editModal').style.display = 'none';
}

document.getElementById('editForm').addEventListener('submit', async function(e) {
    e.preventDefault();
    const wagonId = document.getElementById('edit_wagon_id').value;
    const formData = new FormData(this);
    let arrival = formData.get('arrival_time');
    let globalDeadline = formData.get('departure_time');
    let localDeadline = formData.get('local_departure_time');
    if (arrival) formData.set('arrival_time', normalizeDateTimeForEdit(arrival));
    if (globalDeadline) formData.set('departure_time', normalizeDateTimeForEdit(globalDeadline));
    if (localDeadline) formData.set('local_departure_time', normalizeDateTimeForEdit(localDeadline));
    const response = await fetch(`/edit_wagon/${wagonId}`, {
        method: 'POST',
        body: formData
    });
    const result = await response.json();
    if (result.success) {
        alert(result.message);
        location.reload();
    } else {
        alert('Ошибка: ' + result.message);
    }
});

function formatTimer(d, h, m, s) {
    if (d > 0) return d + 'д ' + String(h).padStart(2,'0') + ':' + String(m).padStart(2,'0') + ':' + String(s).padStart(2,'0');
    return String(h).padStart(2,'0') + ':' + String(m).padStart(2,'0') + ':' + String(s).padStart(2,'0');
}

function formatTimerShort(d, h, m) {
    if (d > 0) return d + 'д ' + String(h).padStart(2,'0') + ':' + String(m).padStart(2,'0');
    return String(h).padStart(2,'0') + ':' + String(m).padStart(2,'0');
}

function formatNegativeTime(seconds) {
    let absSec = Math.abs(seconds);
    let s = absSec % 60;
    let m = Math.floor((absSec % 3600) / 60);
    let h = Math.floor(absSec / 3600);
    let d = Math.floor(h / 24);
    h = h % 24;
    let str = '';
    if (d > 0) str += d + 'д ';
    str += String(h).padStart(2,'0') + ':' + String(m).padStart(2,'0') + ':' + String(s).padStart(2,'0');
    return '-' + str;
}

function updateTooltipTimers() {
    if (!activeWagonId) return;
    const now = new Date();
    const ttLoc = document.getElementById('tt-loc');
    const ttGlob = document.getElementById('tt-glob');
    const locTime = ttLoc.getAttribute('data-time');
    if (locTime) {
        const dep = new Date(locTime);
        const diff = Math.floor((dep - now) / 1000);
        if (diff > 0) {
            let s = diff % 60;
            let m = Math.floor((diff % 3600) / 60);
            let h = Math.floor(diff / 3600);
            let d = Math.floor(h / 24);
            h = h % 24;
            ttLoc.innerHTML = formatTimer(d, h, m, s);
            ttLoc.style.color = diff < 3600 ? '#e67e22' : '#2ecc71';
        } else {
            ttLoc.innerHTML = 'ПРОСРОЧЕНО на ' + formatNegativeTime(diff);
            ttLoc.style.color = '#e74c3c';
        }
    }
    const globTime = ttGlob.getAttribute('data-time');
    if (globTime) {
        const dep = new Date(globTime);
        const diff = Math.floor((dep - now) / 1000);
        if (diff > 0) {
            let s = diff % 60;
            let m = Math.floor((diff % 3600) / 60);
            let h = Math.floor(diff / 3600);
            let d = Math.floor(h / 24);
            h = h % 24;
            ttGlob.innerHTML = formatTimer(d, h, m, s);
            ttGlob.style.color = diff < 3600 ? '#e67e22' : '#2ecc71';
        } else {
            ttGlob.innerHTML = 'ИСТЕК (просрочка ' + formatNegativeTime(diff) + ')';
            ttGlob.style.color = '#e74c3c';
        }
    }
}

function updateAllTimers() {
    const now = new Date();
    document.querySelectorAll('[id^="timer-loc-"]').forEach(el => {
        const timeStr = el.getAttribute('data-time');
        if (timeStr) {
            const dep = new Date(timeStr);
            const diff = Math.floor((dep - now) / 1000);
            if (diff > 0) {
                let s = diff % 60;
                let m = Math.floor((diff % 3600) / 60);
                let h = Math.floor(diff / 3600);
                let d = Math.floor(h / 24);
                h = h % 24;
                el.innerHTML = '<span class="timer-label">⏳</span> ' + formatTimer(d, h, m, s);
                const wagon = el.closest('.wagon');
                if (diff < 300 && wagon && !wagon.classList.contains('active')) {
                    wagon.classList.remove('wagon-normal', 'wagon-return-highlight', 'wagon-warn');
                    wagon.classList.add('wagon-overdue');
                }
            } else {
                el.innerHTML = '<span class="timer-label">⚠️</span> ПРОСРОЧЕНО на ' + formatNegativeTime(diff);
                const wagon = el.closest('.wagon');
                if (wagon && !wagon.classList.contains('active')) {
                    wagon.classList.remove('wagon-normal', 'wagon-return-highlight', 'wagon-warn');
                    wagon.classList.add('wagon-overdue');
                }
            }
        }
    });
    document.querySelectorAll('[id^="timer-glob-"]').forEach(el => {
        const timeStr = el.getAttribute('data-time');
        if (timeStr) {
            const dep = new Date(timeStr);
            const diff = Math.floor((dep - now) / 1000);
            if (diff > 0) {
                let s = diff % 60;
                let m = Math.floor((diff % 3600) / 60);
                let h = Math.floor(diff / 3600);
                let d = Math.floor(h / 24);
                h = h % 24;
                el.innerHTML = '<span class="timer-label">🌍</span> ' + formatTimerShort(d, h, m);
                const wagon = el.closest('.wagon');
                if (diff <= 0 && wagon) {
                    if (wagon.classList.contains('wagon-return-highlight')) {
                        wagon.classList.remove('wagon-return-highlight');
                        wagon.classList.add('wagon-global-overdue');
                    } else if (wagon.classList.contains('wagon-normal')) {
                        wagon.classList.remove('wagon-normal');
                        wagon.classList.add('wagon-global-overdue-normal');
                    }
                }
            } else {
                el.innerHTML = '<span class="timer-label">💀</span> ИСТЕК (просрочка ' + formatNegativeTime(diff) + ')';
                const wagon = el.closest('.wagon');
                if (wagon) {
                    if (wagon.classList.contains('wagon-return-highlight')) {
                        wagon.classList.remove('wagon-return-highlight');
                        wagon.classList.add('wagon-global-overdue');
                    } else if (wagon.classList.contains('wagon-normal')) {
                        wagon.classList.remove('wagon-normal');
                        wagon.classList.add('wagon-global-overdue-normal');
                    }
                }
            }
        }
    });
    updateTooltipTimers();
}

function fetchWagonInfo(num) {
    if (!num) return;
    fetch('/api/wagon_info?num=' + encodeURIComponent(num))
        .then(r => r.json())
        .then(d => {
            if (d.owner) document.getElementById('new-wagon-owner').value = d.owner;
            if (d.org) document.getElementById('new-wagon-org').value = d.org;
        })
        .catch(e => console.log('Ошибка:', e));
}

function updateDashboard() {
    fetch('/api/dashboard_data')
        .then(response => response.json())
        .then(data => {
            const totalCountElem = document.querySelector('.total-count');
            if (totalCountElem) totalCountElem.innerHTML = '📦 Всего: ' + data.total_wagons;
            for (let newTrack of data.tracks) {
                const trackWrapper = document.querySelector(`.track-wrapper[data-track-id="${newTrack.id}"]`);
                if (!trackWrapper) continue;
                const trackHeader = trackWrapper.querySelector('.track-header span:last-child');
                if (trackHeader) trackHeader.innerHTML = '📦 Вагонов: ' + newTrack.wagons.length;
                const trackBody = trackWrapper.querySelector('.track-body');
                if (!trackBody) continue;
                let newHtml = '';
                for (let w of newTrack.wagons) {
                    let leftPct = (w.pos / newTrack.total) * 100;
                    let wagonClass = 'wagon';
                    if (w.loc_overdue) wagonClass += ' wagon-overdue';
                    else if (w.loc_raw < 3600) wagonClass += ' wagon-warn';
                    else if (w.is_highlighted_return && w.is_global_overdue) wagonClass += ' wagon-global-overdue';
                    else if (w.is_highlighted_return) wagonClass += ' wagon-return-highlight';
                    else if (w.is_global_overdue) wagonClass += ' wagon-global-overdue-normal';
                    else wagonClass += ' wagon-normal';
                    
                    let locTimerHtml = '';
                    if (w.loc_iso) {
                        let days = Math.floor(w.loc_raw / 86400);
                        let hours = Math.floor((w.loc_raw % 86400) / 3600);
                        let minutes = Math.floor((w.loc_raw % 3600) / 60);
                        let seconds = Math.floor(w.loc_raw % 60);
                        locTimerHtml = `<div class="wagon-timer" id="timer-loc-${w.id}" data-time="${w.loc_iso}" data-raw="${w.loc_raw}"><span class="timer-label">⏳</span><span>${days}д ${String(hours).padStart(2,'0')}:${String(minutes).padStart(2,'0')}:${String(seconds).padStart(2,'0')}</span></div>`;
                    }
                    let globTimerHtml = '';
                    if (w.glob_iso) {
                        let days = Math.floor(w.glob_raw / 86400);
                        let hours = Math.floor((w.glob_raw % 86400) / 3600);
                        let minutes = Math.floor((w.glob_raw % 3600) / 60);
                        globTimerHtml = `<div class="wagon-timer" id="timer-glob-${w.id}" data-time="${w.glob_iso}" data-raw="${w.glob_raw}"><span class="timer-label">🌍</span><span>${days}д ${String(hours).padStart(2,'0')}:${String(minutes).padStart(2,'0')}</span></div>`;
                    }
                    
                    newHtml += `
                        <div class="${wagonClass}" id="wagon-${w.id}"
                             style="width:130px; left: ${leftPct}%;"
                             data-id="${w.id}"
                             data-num="${w.num.replace(/&/g, '&amp;').replace(/</g, '&lt;').replace(/>/g, '&gt;')}"
                             data-owner="${w.owner.replace(/&/g, '&amp;').replace(/</g, '&lt;').replace(/>/g, '&gt;')}"
                             data-org="${w.org.replace(/&/g, '&amp;').replace(/</g, '&lt;').replace(/>/g, '&gt;')}"
                             data-note="${w.note.replace(/&/g, '&amp;').replace(/</g, '&lt;').replace(/>/g, '&gt;')}"
                             data-arrival="${w.arrival}"
                             data-loc-iso="${w.loc_iso}"
                             data-glob-iso="${w.glob_iso}"
                             data-track-name="${newTrack.name.replace(/&/g, '&amp;').replace(/</g, '&lt;').replace(/>/g, '&gt;')}"
                             data-loc-raw="${w.loc_raw}"
                             data-glob-raw="${w.glob_raw}">
                            <div class="wagon-number">${w.num.replace(/&/g, '&amp;').replace(/</g, '&lt;').replace(/>/g, '&gt;')}</div>
                            ${locTimerHtml}
                            ${globTimerHtml}
                        </div>
                    `;
                }
                trackBody.innerHTML = newHtml;
            }
            updateAllTimers();
        })
        .catch(err => console.log('Ошибка обновления дашборда:', err));
}

// Единый обработчик кликов для открытия и закрытия тултипа (делегирование)
document.addEventListener('click', function(e) {
    const wagon = e.target.closest('.wagon');
    if (wagon) {
        e.stopPropagation();
        const id = wagon.dataset.id;
        const num = wagon.dataset.num;
        const owner = wagon.dataset.owner;
        const org = wagon.dataset.org;
        const note = wagon.dataset.note;
        const arrival = wagon.dataset.arrival;
        const locIso = wagon.dataset.locIso;
        const globIso = wagon.dataset.globIso;
        const trackName = wagon.dataset.trackName;
        const locRaw = parseFloat(wagon.dataset.locRaw);
        const globRaw = parseFloat(wagon.dataset.globRaw);
        openTooltip(wagon, id, num, owner, org, note, arrival, locIso, globIso, trackName, locRaw, globRaw);
    } else {
        closeTooltip();
    }
});

setInterval(updateAllTimers, 1000);
setInterval(updateDashboard, 5000);
</script>
</body>
</html>"""

HISTORY_SPOILER_TEMPLATE = """<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8">
    <title>{{ title }}</title>
    <style>
        body { font-family: Arial, sans-serif; padding: 20px; background: #f4f6f9; }
        .container { max-width: 1200px; margin: 0 auto; background: white; padding: 20px; border-radius: 10px; }
        h1 { color: #2c3e50; border-bottom: 2px solid #3498db; padding-bottom: 10px; }
        .nav-links { display: flex; gap: 10px; margin-bottom: 20px; flex-wrap: wrap; }
        .nav-link { display: inline-block; padding: 10px 20px; background: #3498db; color: white; text-decoration: none; border-radius: 5px; }
        .nav-link-archive { background: #7f8c8d; }
        .nav-link-excel { background: #217346; }
        .nav-link-help { background: #16a085; }
        .search-box { margin-bottom: 20px; display: flex; gap: 10px; align-items: center; }
        .search-box input { flex: 1; padding: 10px; font-size: 16px; border: 1px solid #bdc3c7; border-radius: 5px; }
        .search-box button { padding: 10px 15px; background: #3498db; color: white; border: none; border-radius: 5px; cursor: pointer; }
        .search-box button:hover { background: #2980b9; }
        .no-results { text-align: center; padding: 20px; color: #e74c3c; font-weight: bold; display: none; }
        details { border: 1px solid #ddd; border-radius: 6px; margin-bottom: 10px; background: white; }
        summary {
            padding: 12px;
            cursor: pointer;
            background: #ecf0f1;
            font-weight: bold;
            display: grid;
            grid-template-columns: 150px 120px 1fr 100px;
            align-items: center;
            gap: 10px;
            list-style: none;
        }
        summary::-webkit-details-marker { display: none; }
        summary::marker { display: none; }
        .wagon-num { font-size: 1.2em; color: #2980b9; }
        .status-added { background: #27ae60; color: white; padding: 2px 8px; border-radius: 12px; display: inline-block; text-align: center; }
        .status-moved { background: #f39c12; color: white; padding: 2px 8px; border-radius: 12px; display: inline-block; text-align: center; }
        .status-departed { background: #e74c3c; color: white; padding: 2px 8px; border-radius: 12px; display: inline-block; text-align: center; }
        .status-edit { background: #8e44ad; color: white; padding: 2px 8px; border-radius: 12px; display: inline-block; text-align: center; }
        .excel-wagon-btn {
            background: #217346;
            color: white;
            padding: 4px 10px;
            border-radius: 4px;
            text-decoration: none;
            font-size: 12px;
            display: inline-flex;
            align-items: center;
            gap: 4px;
            justify-self: end;
            white-space: nowrap;
        }
        .excel-wagon-btn:hover { background: #1e5c3a; }
        .wagon-info { justify-self: start; }
        .details-content { padding: 15px; border-top: 1px solid #ddd; overflow-x: auto; }
        .history-table {
            width: 100%;
            border-collapse: collapse;
            font-size: 14px;
            margin-top: 5px;
        }
        .history-table th, .history-table td {
            border: 1px solid #ddd;
            padding: 8px;
            text-align: left;
            vertical-align: top;
        }
        .history-table th {
            background-color: #34495e;
            color: white;
            font-weight: bold;
        }
        .history-table tr:nth-child(even) {
            background-color: #f9f9f9;
        }
        .history-table tr:hover {
            background-color: #f1f1f1;
        }
        .history-table th:nth-child(1), .history-table td:nth-child(1) { width: 100px; }
        .history-table th:nth-child(2), .history-table td:nth-child(2) { width: 120px; }
        .history-table th:nth-child(3), .history-table td:nth-child(3) { width: 120px; }
        .history-table th:nth-child(4), .history-table td:nth-child(4) { width: 150px; }
        .history-table th:nth-child(5), .history-table td:nth-child(5) { width: 150px; }
        .history-table th:nth-child(6), .history-table td:nth-child(6) { width: 250px; word-break: break-word; white-space: normal; }
        .history-table th:nth-child(7), .history-table td:nth-child(7) { width: 160px; }
        .edit-history-btn { background: none; border: none; cursor: pointer; color: #2980b9; font-size: 14px; margin-left: 8px; }
        .edit-history-btn:hover { color: #e67e22; }
        .edit-history-locked { color: #95a5a6; font-size: 12px; margin-left: 8px; cursor: help; }
        .wagon-details { display: block; }
    </style>
</head>
<body>
<div class="container">
    <div class="nav-links">
        <a href="/" class="nav-link">🏠 На главную</a>
        <a href="/archive" class="nav-link nav-link-archive">🗄️ Архив</a>
        <a href="/export_history_excel" class="nav-link nav-link-excel">📊 Скачать Excel (все вагоны)</a>
        <a href="/help" class="nav-link nav-link-help">❓ Помощь</a>
    </div>
    <h1>{{ title }}</h1>

    <div class="search-box">
        <input type="text" id="searchInput" placeholder="🔍 Поиск по номеру вагона..." autocomplete="off">
        <button id="clearSearch">Очистить</button>
    </div>
    <div id="noResultsMsg" class="no-results">❌ Ничего не найдено</div>

    <div id="historyContainer">
    {% if not history_groups %}
        <p>История пуста.</p>
    {% else %}
        {% for item in history_groups %}
            <details class="wagon-details" data-wagon-num="{{ item.num|lower }}">
                <summary>
                    <span class="wagon-num">{{ item.num }}</span>
                    <span class="status-{% if 'Добавлен' in item.last_status %}added{% elif 'Перемещен' in item.last_status %}moved{% elif 'Изменён' in item.last_status %}edit{% else %}departed{% endif %}">{{ item.last_status|safe }}</span>
                    <span class="wagon-info">📅 {{ item.last_time[:16] }} | Событий: {{ item.count }}</span>
                    <a href="/export_wagon_history/{{ item.num }}" class="excel-wagon-btn" title="Выгрузить историю только этого вагона">📊 Excel</a>
                </summary>
                <div class="details-content">
                    <table class="history-table">
                        <thead>
                            <tr><th>Действие</th><th>Откуда</th><th>Куда</th><th>ТК</th><th>Орг</th><th>Примечание</th><th>Время</th></tr>
                        </thead>
                        <tbody>
                            {% for ev in item.events %}
                                <tr>
                                    <td>{{ ev.action|safe }}</td>
                                    <td>{{ ev.from }}</td>
                                    <td>{{ ev.to }}</td>
                                    <td>{{ ev.owner }}</td>
                                    <td>{{ ev.org }}</td>
                                    <td>{{ ev.note }}</td>
                                    <td>
                                        {{ ev.time }}
                                        {% if session_role in ('supervisor', 'admin') %}
                                            {% if ev.is_last %}
                                                <button class="edit-history-btn" data-id="{{ ev.id }}" data-time="{{ ev.time }}" title="Редактировать дату последнего события">✏️</button>
                                            {% else %}
                                                <span class="edit-history-locked" title="Редактирование доступно только для последнего события, чтобы не нарушать хронологию">🔒</span>
                                            {% endif %}
                                        {% endif %}
                                    </td>
                                </tr>
                            {% endfor %}
                        </tbody>
                    </table>
                </div>
            </details>
        {% endfor %}
    {% endif %}
    </div>
</div>
<script>
function filterWagons() {
    const searchTerm = document.getElementById('searchInput').value.trim().toLowerCase();
    const containers = document.querySelectorAll('.wagon-details');
    let visibleCount = 0;
    containers.forEach(container => {
        const wagonNum = container.getAttribute('data-wagon-num');
        if (wagonNum && wagonNum.includes(searchTerm)) {
            container.style.display = '';
            visibleCount++;
        } else {
            container.style.display = 'none';
        }
    });
    const noResultsDiv = document.getElementById('noResultsMsg');
    if (visibleCount === 0 && searchTerm !== '') {
        noResultsDiv.style.display = 'block';
    } else {
        noResultsDiv.style.display = 'none';
    }
}
document.getElementById('searchInput').addEventListener('input', filterWagons);
document.getElementById('clearSearch').addEventListener('click', function() {
    document.getElementById('searchInput').value = '';
    filterWagons();
});
document.querySelectorAll('.edit-history-btn').forEach(btn => {
    btn.addEventListener('click', function(e) {
        e.stopPropagation();
        const historyId = this.dataset.id;
        const oldTime = this.dataset.time;
        const newTime = prompt('Введите новую дату и время (поддерживается ГГГГ-ММ-ДД ЧЧ:ММ, ДД.ММ.ГГГГ ЧЧ:ММ, ДДММГГГГ и т.д.):', oldTime);
        if (newTime && newTime !== oldTime) {
            fetch(`/edit_history/${historyId}`, {
                method: 'POST',
                headers: { 'Content-Type': 'application/x-www-form-urlencoded' },
                body: 'timestamp=' + encodeURIComponent(newTime)
            })
            .then(res => res.json())
            .then(data => {
                if (data.success) {
                    alert(data.message);
                    location.reload();
                } else {
                    alert('Ошибка: ' + data.error);
                }
            })
            .catch(err => alert('Ошибка запроса: ' + err));
        }
    });
});
</script>
</body>
</html>"""

ARCHIVE_SPOILER_TEMPLATE = """<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8">
    <title>Архив</title>
    <style>
        body { font-family: Arial, sans-serif; padding: 20px; background: #f4f6f9; }
        .container { max-width: 1000px; margin: 0 auto; background: white; padding: 20px; border-radius: 10px; }
        h1 { color: #2c3e50; border-bottom: 2px solid #7f8c8d; padding-bottom: 10px; }
        .nav-links { display: flex; gap: 10px; margin-bottom: 20px; flex-wrap: wrap; }
        .nav-link { display: inline-block; padding: 10px 20px; background: #7f8c8d; color: white; text-decoration: none; border-radius: 5px; }
        .nav-link-history { background: #3498db; }
        .nav-link-excel { background: #217346; }
        .nav-link-help { background: #16a085; }
        .search-box { margin-bottom: 20px; display: flex; gap: 10px; align-items: center; }
        .search-box input { flex: 1; padding: 10px; font-size: 16px; border: 1px solid #bdc3c7; border-radius: 5px; }
        .search-box button { padding: 10px 15px; background: #3498db; color: white; border: none; border-radius: 5px; cursor: pointer; }
        .search-box button:hover { background: #2980b9; }
        .no-results { text-align: center; padding: 20px; color: #e74c3c; font-weight: bold; display: none; }
        details { border: 1px solid #ddd; border-radius: 6px; margin-bottom: 10px; background: white; }
        summary {
            padding: 12px;
            cursor: pointer;
            background: #ecf0f1;
            font-weight: bold;
            display: grid;
            grid-template-columns: 150px 80px 200px 1fr 100px;
            align-items: center;
            gap: 10px;
            list-style: none;
        }
        summary::-webkit-details-marker { display: none; }
        summary::marker { display: none; }
        .wagon-num { font-size: 1.2em; color: #2980b9; }
        .wagon-status { background: #e74c3c; color: white; padding: 2px 8px; border-radius: 12px; display: inline-block; text-align: center; }
        .excel-wagon-btn {
            background: #217346;
            color: white;
            padding: 4px 10px;
            border-radius: 4px;
            text-decoration: none;
            font-size: 12px;
            display: inline-flex;
            align-items: center;
            gap: 4px;
            justify-self: end;
            white-space: nowrap;
        }
        .excel-wagon-btn:hover { background: #1e5c3a; }
        .details-content { padding: 15px; border-top: 1px solid #ddd; overflow-x: auto; }
        .history-table {
            width: 100%;
            border-collapse: collapse;
            font-size: 14px;
            margin-top: 5px;
        }
        .history-table th, .history-table td {
            border: 1px solid #ddd;
            padding: 8px;
            text-align: left;
            vertical-align: top;
        }
        .history-table th {
            background-color: #34495e;
            color: white;
            font-weight: bold;
        }
        .history-table tr:nth-child(even) {
            background-color: #f9f9f9;
        }
        .history-table tr:hover {
            background-color: #f1f1f1;
        }
        .history-table th:nth-child(1), .history-table td:nth-child(1) { width: 100px; }
        .history-table th:nth-child(2), .history-table td:nth-child(2) { width: 120px; }
        .history-table th:nth-child(3), .history-table td:nth-child(3) { width: 120px; }
        .history-table th:nth-child(4), .history-table td:nth-child(4) { width: 250px; word-break: break-word; white-space: normal; }
        .history-table th:nth-child(5), .history-table td:nth-child(5) { width: 160px; }
        .wagon-details { display: block; }
    </style>
</head>
<body>
<div class="container">
    <div class="nav-links">
        <a href="/" class="nav-link">🏠 На главную</a>
        <a href="/history" class="nav-link nav-link-history">📜 История</a>
        <a href="/export_archive_excel" class="nav-link nav-link-excel">📊 Скачать Excel (все вагоны)</a>
        <a href="/help" class="nav-link nav-link-help">❓ Помощь</a>
    </div>
    <h1>🗄️ Архив</h1>

    <div class="search-box">
        <input type="text" id="searchInput" placeholder="🔍 Поиск по номеру вагона..." autocomplete="off">
        <button id="clearSearch">Очистить</button>
    </div>
    <div id="noResultsMsg" class="no-results">❌ Ничего не найдено</div>

    <div id="archiveContainer">
    {% if not archive_groups %}
        <p>Архив пуст.</p>
    {% else %}
        {% for item in archive_groups %}
            <details class="wagon-details" data-wagon-num="{{ item.num|lower }}">
                <summary>
                    <span class="wagon-num">{{ item.num }}</span>
                    <span class="wagon-status">Убыл</span>
                    <span>{{ item.owner }}</span>
                    <span>{{ item.org }}</span>
                    <span>📅 {{ item.dep[:16] }}</span>
                    <span>Событий: {{ item.count }}</span>
                    <a href="/export_wagon_archive/{{ item.num }}" class="excel-wagon-btn" title="Выгрузить историю только этого вагона">📊 Excel</a>
                </summary>
                <div class="details-content">
                    <table class="history-table">
                        <thead>
                            <tr><th>Действие</th><th>Откуда</th><th>Куда</th><th>Примечание</th><th>Время</th></tr>
                        </thead>
                        <tbody>
                            {% for ev in item.events %}
                            <tr>
                                <td>{{ ev.action|safe }}</td>
                                <td>{{ ev.from }}</td>
                                <td>{{ ev.to }}</td>
                                <td>{{ ev.note }}</td>
                                <td>{{ ev.time }}</td>
                            </tr>
                            {% endfor %}
                        </tbody>
                    </table>
                </div>
            </details>
        {% endfor %}
    {% endif %}
    </div>
</div>
<script>
function filterWagons() {
    const searchTerm = document.getElementById('searchInput').value.trim().toLowerCase();
    const containers = document.querySelectorAll('.wagon-details');
    let visibleCount = 0;
    containers.forEach(container => {
        const wagonNum = container.getAttribute('data-wagon-num');
        if (wagonNum && wagonNum.includes(searchTerm)) {
            container.style.display = '';
            visibleCount++;
        } else {
            container.style.display = 'none';
        }
    });
    const noResultsDiv = document.getElementById('noResultsMsg');
    if (visibleCount === 0 && searchTerm !== '') {
        noResultsDiv.style.display = 'block';
    } else {
        noResultsDiv.style.display = 'none';
    }
}
document.getElementById('searchInput').addEventListener('input', filterWagons);
document.getElementById('clearSearch').addEventListener('click', function() {
    document.getElementById('searchInput').value = '';
    filterWagons();
});
</script>
</body>
</html>"""

HELP_TEMPLATE = """<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8">
    <title>Инструкция диспетчера - ЖД Диспетчерская</title>
    <style>
        body { font-family: 'Segoe UI', Arial, sans-serif; background: #f0f2f5; margin: 0; padding: 20px; }
        .container { max-width: 1000px; margin: 0 auto; background: white; padding: 30px; border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }
        h1 { color: #2c3e50; border-bottom: 2px solid #3498db; padding-bottom: 10px; }
        h2 { color: #2980b9; margin-top: 25px; }
        .nav-link { display: inline-block; padding: 8px 15px; background: #3498db; color: white; text-decoration: none; border-radius: 5px; margin-bottom: 20px; }
        .color-box { display: inline-block; width: 20px; height: 20px; border-radius: 4px; vertical-align: middle; margin-right: 8px; }
        table { width: 100%; border-collapse: collapse; margin: 15px 0; }
        th, td { padding: 10px; text-align: left; border-bottom: 1px solid #ddd; }
        th { background: #ecf0f1; }
        .note { background: #fef9e7; padding: 10px; border-left: 4px solid #f39c12; margin: 15px 0; }
        .address { background: #e8f4fd; padding: 12px; border-radius: 6px; font-family: monospace; font-size: 1.2em; text-align: center; margin: 20px 0; }
    </style>
</head>
<body>
<div class="container">
    <a href="/" class="nav-link">🏠 На главную</a>
    <h1>🚂 Инструкция для диспетчера</h1>
    
    <h2>1. Как подключиться</h2>
    <p>Откройте браузер (Chrome, Firefox, Edge) и введите адрес:</p>
    <div class="address">http://{{ server_ip }}:5000</div>
    <p>Если вы работаете с того же компьютера, где запущена программа, используйте <strong>http://127.0.0.1:5000</strong>.</p>
    <div class="note">📌 После входа вы увидите главный экран со списком путей и вагонов.</div>
    
    <h2>2. Главный экран</h2>
    <p>На главной странице отображаются:</p>
    <ul>
        <li><strong>Пути</strong> (станции, посты, цеха) в виде горизонтальных полос.</li>
        <li><strong>Вагоны</strong> – цветные прямоугольники с номером и таймерами.</li>
        <li>Панель управления слева – <strong>«Новый вагон»</strong>, справа – <strong>«Переместить»</strong>.</li>
        <li>Вверху – кнопки: История, Архив, Excel, Помощь (эта страница).</li>
    </ul>
    
    <h2>3. Добавление нового вагона</h2>
    <ul>
        <li>Заполните поля: <strong>№ вагона</strong>, <strong>Транспортная компания</strong>, <strong>Организация</strong>.</li>
        <li>При желании укажите <strong>глобальный срок</strong> – общее время на станции. Для этого задайте дни, часы, минуты.</li>
        <li>Можно указать <strong>начало отсчёта</strong> (дата и время). Если оставить поля пустыми – срок начнётся с текущего момента.</li>
        <li>Выберите <strong>путь</strong>, на который ставится вагон.</li>
        <li>Нажмите <strong>«Распределить»</strong>.</li>
    </ul>
    <div class="note">💡 Если вагон уже был в архиве, он восстановится с сохранением истории.</div>
    
    <h2>4. Перемещение вагона</h2>
    <ul>
        <li>В правой панели найдите вагон (помогает поле <strong>«Поиск вагона»</strong>).</li>
        <li>Выберите его из списка – примечание подставится автоматически.</li>
        <li>При необходимости измените <strong>примечание</strong>.</li>
        <li>Задайте <strong>локальный срок</strong> – время, которое вагон должен провести на новом пути (дни, часы, минуты).</li>
        <li>Укажите <strong>начало отсчёта</strong> (дата/время) или оставьте пустым – будет использовано текущее время.</li>
        <li>Выберите <strong>путь назначения</strong> и нажмите <strong>«Переместить»</strong>.</li>
    </ul>
    
    <h2>5. Цвета вагонов (что означают)</h2>
    <table>
        <tr><th>Цвет</th><th>Значение</th></tr>
        <tr><td><span class="color-box" style="background: linear-gradient(135deg,#3498db,#2980b9);"></span> Синий</span><th>Обычный вагон (срок в норме или не задан)</th></tr>
        <tr><td><span class="color-box" style="background: linear-gradient(135deg,#8e44ad,#9b59b6); border:1px solid #f1c40f;"></span> Фиолетовый (с жёлтой рамкой)</span><th>Вагон уже побывал на возвратном пути («Пост №2» или «Ст. Черкасов Камень»)</th></tr>
        <tr><td><span class="color-box" style="background: linear-gradient(135deg,#6c3483,#4a235a); border:2px solid #e74c3c;"></span> Тёмно-фиолетовый</span><th>Был на возвратном пути И истёк глобальный срок</th></tr>
        <tr><td><span class="color-box" style="background: linear-gradient(135deg,#922b21,#641e16); border:2px solid #e74c3c;"></span> Тёмно-красный</span><th>Обычный вагон с истёкшим глобальным сроком</th></tr>
        <tr><td><span class="color-box" style="background: linear-gradient(135deg,#f39c12,#d35400);"></span> Оранжевый</span><th>Локальный срок истекает менее чем через час</th></tr>
        <tr><td><span class="color-box" style="background: linear-gradient(135deg,#e74c3c,#c0392b);"></span> Красный</span><th>Локальный срок уже истёк (просрочка)</th></tr>
    </table>
    <p><strong>Локальный срок</strong> – время на текущем пути. <strong>Глобальный срок</strong> – общее время на станции.</p>
    
    <h2>6. Информация о вагоне</h2>
    <p>Кликните по вагону – появится всплывающее окно с подробностями:</p>
    <ul>
        <li>Номер вагона, транспортная компания, организация.</li>
        <li>Примечание, время прибытия.</li>
        <li><strong>Таймеры</strong> локального и глобального сроков (обратный отсчёт).</li>
        <li>Если вагон находится на возвратном пути («Пост №2» или «Ст. Черкасов Камень»), будет кнопка <strong>«УБРАТЬ В АРХИВ»</strong>.</li>
    </ul>
    <p>Чтобы закрыть окно, нажмите «Закрыть» или кликните вне его.</p>
    
    <h2>7. История и архив</h2>
    <p>Кнопки <strong>«История»</strong> и <strong>«Архив»</strong> вверху страницы.</p>
    <ul>
        <li><strong>История</strong> – все перемещения активных вагонов (сгруппированы по вагонам, раскрывающиеся блоки).</li>
        <li><strong>Архив</strong> – вагоны, которые были убраны с возвратных путей. Также сгруппированы с полной историей.</li>
    </ul>
    <p>Внутри каждого блока есть кнопка <strong>«📊 Excel»</strong> – выгружает историю конкретного вагона в отдельный файл.</p>
    
    <h2>8. Выгрузка в Excel</h2>
    <ul>
        <li>На главной странице – кнопка <strong>«Excel»</strong> (отчёт по всем активным вагонам).</li>
        <li>В разделе «История» – кнопка <strong>«Скачать Excel (все вагоны)»</strong> (полная история перемещений).</li>
        <li>В разделе «Архив» – кнопка <strong>«Скачать Excel (все вагоны)»</strong> (сводка + детализация по архиву).</li>
        <li>Для каждого отдельного вагона (как в истории, так и в архиве) есть своя кнопка Excel.</li>
    </ul>
    
    <h2>9. Редактирование дат в истории (для ролей supervisor и admin)</h2>
    <p>На странице «История» у последнего события каждого вагона появляется кнопка <strong>✏️</strong>.</p>
    <p><strong>Почему только последнее событие?</strong> Изменение даты более раннего события нарушило бы хронологию перемещений и могло бы привести к неверной работе таймеров локального срока. Если вы ошиблись в дате не последнего события, рекомендуется отредактировать время прибытия или сроки вагона (кнопка «Редактировать вагон»), либо удалить и добавить вагон заново.</p>
    <p>При изменении даты последнего события проверяется, что новая дата не нарушает хронологию, и автоматически пересчитывается локальный срок (если был задан). Изменение записывается в журнал действий.</p>
    
    <h2>10. Завершение работы</h2>
    <p>Диспетчеру достаточно закрыть вкладку браузера. Программа на сервере продолжит работать. Если нужно полностью остановить сервер – это делает администратор на серверной машине.</p>
    
    <div class="note">📞 При возникновении проблем обратитесь к системному администратору.</div>
    <p style="text-align: center; margin-top: 30px;">© ЖД Диспетчерская, АО "Знамя"<br>Версия {{ version }} (сетевая, для диспетчера)</p>
</div>
</body>
</html>"""

# ==================== АДМИНИСТРАТИВНЫЕ МАРШРУТЫ ====================
@app.route('/admin/backup', methods=['POST'])
def create_backup():
    if request.user_role != 'admin':
        return "Доступ запрещён", 403
    try:
        date_str = datetime.now().strftime('%Y%m%d')
        daily_dir = os.path.join(BACKUP_DIR, date_str)
        if not os.path.exists(daily_dir):
            os.makedirs(daily_dir)
        time_str = datetime.now().strftime('%H%M%S')
        backup_name = f"rail_yard_backup_{date_str}_{time_str}.db"
        backup_path = os.path.join(daily_dir, backup_name)
        shutil.copy2(DB_NAME, backup_path)
        log_action('backup_create', details=f"Создана копия: {backup_path}")
        return f"✅ Резервная копия создана: {backup_path}", 200
    except Exception as e:
        return f"❌ Ошибка создания бэкапа: {str(e)}", 500

@app.route('/admin/backups')
def list_backups():
    if request.user_role != 'admin':
        return "Доступ запрещён", 403
    all_backups = glob.glob(os.path.join(BACKUP_DIR, '**', '*.db'), recursive=True)
    backups_info = []
    for path in all_backups:
        stat = os.stat(path)
        rel_path = os.path.relpath(path, BACKUP_DIR).replace('\\', '/')
        backups_info.append({
            'name': os.path.basename(path),
            'rel_path': rel_path,
            'size': stat.st_size,
            'modified': datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S')
        })
    backups_info.sort(key=lambda x: x['modified'], reverse=True)
    html = """
    <html>
    <head><title>Резервные копии</title>
    <style>
        body { font-family: monospace; padding: 20px; background: #f0f2f5; }
        table { border-collapse: collapse; width: 100%; background: white; }
        th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }
        th { background: #34495e; color: white; }
        .btn { display: inline-block; padding: 5px 10px; background: #3498db; color: white; text-decoration: none; border-radius: 4px; margin-right: 5px; }
        .btn-danger { background: #e74c3c; }
        .nav-bar { margin-bottom: 15px; }
    </style>
    </head>
    <body>
    <h1>📦 Резервные копии базы данных</h1>
    <div class="nav-bar">
        <a href="/" class="btn">🏠 На главную</a>
        <a href="/admin/logs" class="btn">📜 Журнал действий</a>
        <a href="/admin/ip_users" class="btn">🔗 Привязка IP</a>
        <a href="/changelog" class="btn">📋 Список изменений</a>
    </div>
    <p><a href="#" onclick="event.preventDefault(); fetch('/admin/backup', {method:'POST'}).then(r=>r.text()).then(alert);" class="btn">➕ Создать новую копию</a></p>
    <table>
        <tr><th>Имя файла</th><th>Размер (КБ)</th><th>Дата изменения</th><th>Действия</th></tr>
    """
    for b in backups_info:
        size_kb = b['size'] / 1024
        html += f"""
        <tr>
            <td>{b['name']}</td>
            <td>{size_kb:.1f}</td>
            <td>{b['modified']}</td>
            <td>
                <a href="/admin/download_backup?rel_path={b['rel_path']}" class="btn">📥 Скачать</a>
                <a href="#" onclick="restore('{b['rel_path']}')" class="btn btn-danger">🔄 Восстановить</a>
            </td>
        </tr>
        """
    html += """
    </table>
    <script>
    function restore(relPath) {
        if (confirm('Восстановление заменит текущую базу данных! Сделайте резервную копию перед восстановлением. Продолжить?')) {
            fetch('/admin/restore', {
                method: 'POST',
                headers: {'Content-Type': 'application/x-www-form-urlencoded'},
                body: 'rel_path=' + encodeURIComponent(relPath)
            }).then(r => r.text()).then(alert).then(() => location.reload());
        }
    }
    </script>
    </body>
    </html>
    """
    return html

@app.route('/admin/download_backup')
def download_backup():
    if request.user_role != 'admin':
        return "Доступ запрещён", 403
    rel_path = request.args.get('rel_path')
    if not rel_path:
        return "Не указан путь", 400
    full_path = os.path.abspath(os.path.join(BACKUP_DIR, rel_path))
    if not full_path.startswith(os.path.abspath(BACKUP_DIR)):
        return "Неверный путь", 403
    if not os.path.exists(full_path):
        return "Файл не найден", 404
    return send_file(full_path, as_attachment=True, download_name=os.path.basename(full_path))

@app.route('/admin/restore', methods=['POST'])
def restore_backup():
    if request.user_role != 'admin':
        return "Доступ запрещён", 403
    rel_path = request.form.get('rel_path')
    if not rel_path:
        return "Не указан путь", 400
    full_path = os.path.abspath(os.path.join(BACKUP_DIR, rel_path))
    if not full_path.startswith(os.path.abspath(BACKUP_DIR)):
        return "Неверный путь", 403
    if not os.path.exists(full_path):
        return f"Файл не найден: {full_path}", 404
    try:
        temp_backup = os.path.join(BACKUP_DIR, f"pre_restore_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db")
        shutil.copy2(DB_NAME, temp_backup)
        shutil.copy2(full_path, DB_NAME)
        log_action('backup_restore', details=f"Восстановлена БД из {rel_path}")
        return f"✅ База данных восстановлена из {rel_path}. Рекомендуется перезапустить программу."
    except Exception as e:
        return f"❌ Ошибка восстановления: {str(e)}", 500

@app.route('/admin/logs')
def view_logs():
    if request.user_role != 'admin':
        return "Доступ запрещён", 403
    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT * FROM action_log ORDER BY timestamp DESC LIMIT 500")
    logs = c.fetchall()
    conn.close()
    action_translation = {
        'add': 'Добавление вагона',
        'move': 'Перемещение',
        'depart': 'Архивация',
        'backup_create': 'Создание бэкапа',
        'backup_restore': 'Восстановление из бэкапа',
        'ip_user_edit': 'Изменение привязки IP',
        'ip_user_delete': 'Удаление привязки IP',
        'edit': 'Редактирование вагона',
        'backup_auto': 'Автоматический бэкап',
        'edit_history': 'Редактирование истории'
    }
    html = """
    <html>
    <head>
        <title>Журнал действий</title>
        <style>
            body { font-family: monospace; padding: 20px; background: #f0f2f5; }
            table { border-collapse: collapse; width: 100%; background: white; }
            th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }
            th { background: #34495e; color: white; }
            .btn { display: inline-block; padding: 5px 10px; background: #3498db; color: white; text-decoration: none; border-radius: 4px; }
            .btn-excel { background: #27ae60; }
        </style>
    </head>
    <body>
        <h1>📋 Журнал действий</h1>
        <p>
            <a href="/" class="btn">🏠 На главную</a>
            <a href="/export_logs_excel" class="btn btn-excel">📊 Выгрузить в Excel</a>
        </p>
        <table>
            <thead>
                <tr>
                    <th>Время</th>
                    <th>Пользователь</th>
                    <th>IP</th>
                    <th>Действие</th>
                    <th>Вагон</th>
                    <th>Детали</th>
                    <th>Старое</th>
                    <th>Новое</th>
                </tr>
            </thead>
            <tbody>
    """
    for log in logs:
        action_rus = action_translation.get(log[4], log[4])
        html += f"""
                <tr>
                    <td>{log[1]}</td>
                    <td>{log[2]}</td>
                    <td>{log[3]}</td>
                    <td>{action_rus}</td>
                    <td>{log[5] if log[5] else ''}</td>
                    <td>{log[6] if log[6] else ''}</td>
                    <td>{log[7] if log[7] else ''}</td>
                    <td>{log[8] if log[8] else ''}</td>
                </tr>
        """
    html += """
            </tbody>
        </table>
    </body>
    </html>
    """
    return html

@app.route('/export_logs_excel')
def export_logs_excel():
    if request.user_role != 'admin':
        return "Доступ запрещён", 403
    conn = get_conn()
    df = pd.read_sql_query("""
        SELECT 
            timestamp as "Время",
            username as "Пользователь",
            ip_address as "IP-адрес",
            action as "Действие",
            wagon_number as "Номер вагона",
            details as "Детали",
            old_value as "Старое значение",
            new_value as "Новое значение"
        FROM action_log
        ORDER BY timestamp DESC
    """, conn)
    conn.close()
    action_map = {
        'add': 'Добавление вагона',
        'move': 'Перемещение',
        'depart': 'Архивация',
        'backup_create': 'Создание бэкапа',
        'backup_restore': 'Восстановление из бэкапа',
        'ip_user_edit': 'Изменение привязки IP',
        'ip_user_delete': 'Удаление привязки IP',
        'edit': 'Редактирование вагона',
        'backup_auto': 'Автоматический бэкап',
        'edit_history': 'Редактирование истории'
    }
    df['Действие'] = df['Действие'].map(action_map).fillna(df['Действие'])
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Журнал действий', index=False)
        worksheet = writer.sheets['Журнал действий']
        for column in worksheet.columns:
            max_length = 0
            col_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[col_letter].width = adjusted_width
        worksheet.auto_filter.ref = worksheet.dimensions
        from openpyxl.styles import Font, Alignment
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                if cell.column_letter in ('F', 'G', 'H'):
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
    output.seek(0)
    return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name=f"ActionLog_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

@app.route('/admin/ip_users', methods=['GET', 'POST'])
def manage_ip_users():
    if request.user_role != 'admin':
        return "Доступ запрещён", 403
    conn = get_conn()
    c = conn.cursor()
    if request.method == 'POST':
        if request.form.get('delete_ip'):
            del_ip = request.form.get('delete_ip')
            c.execute("DELETE FROM ip_users WHERE ip_address = ?", (del_ip,))
            conn.commit()
            flash(f"Привязка для IP {del_ip} удалена", 'success')
            log_action('ip_user_delete', details=f"Удалена привязка IP {del_ip}")
            conn.close()
            return redirect(url_for('manage_ip_users'))
        ip = request.form.get('ip_address', '').strip()
        username = request.form.get('username', '').strip()
        note = request.form.get('note', '').strip()
        access_allowed = 1 if request.form.get('access_allowed') else 0
        role = request.form.get('role', 'viewer')
        if ip and username:
            c.execute("INSERT OR REPLACE INTO ip_users (ip_address, username, note, access_allowed, role) VALUES (?, ?, ?, ?, ?)", 
                      (ip, username, note, access_allowed, role))
            conn.commit()
            flash(f"Привязка для IP {ip} сохранена", 'success')
            log_action('ip_user_edit', details=f"Добавлен/изменён IP {ip} → {username}, доступ={access_allowed}, роль={role}")
        conn.close()
        return redirect(url_for('manage_ip_users'))
    c.execute("SELECT ip_address, username, note, access_allowed, role FROM ip_users ORDER BY ip_address")
    rows = c.fetchall()
    conn.close()
    
    html = """
    <!DOCTYPE html>
    <html lang="ru">
    <head>
        <meta charset="UTF-8">
        <title>Привязка IP к пользователям</title>
        <style>
            body { font-family: 'Segoe UI', Arial, sans-serif; padding: 20px; background: #f0f2f5; margin: 0; }
            .container { max-width: 1200px; margin: 0 auto; background: white; padding: 20px; border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }
            h1 { color: #2c3e50; margin-top: 0; }
            .btn { display: inline-block; padding: 8px 15px; background: #3498db; color: white; text-decoration: none; border-radius: 5px; margin-right: 5px; font-size: 14px; }
            .btn-excel { background: #27ae60; }
            form { background: #ecf0f1; padding: 20px; border-radius: 8px; margin-bottom: 30px; }
            input, select { padding: 8px; margin: 5px 5px 5px 0; border: 1px solid #bdc3c7; border-radius: 4px; }
            label { margin-right: 10px; }
            table { width: 100%; border-collapse: collapse; background: white; margin-top: 20px; }
            th, td { border: 1px solid #ddd; padding: 10px; text-align: left; vertical-align: top; }
            th { background: #34495e; color: white; font-weight: bold; }
            tr:nth-child(even) { background: #f9f9f9; }
            .delete-btn { background: #e74c3c; color: white; border: none; padding: 5px 10px; border-radius: 4px; cursor: pointer; }
            .nav-bar { margin-bottom: 20px; }
        </style>
    </head>
    <body>
    <div class="container">
        <h1>🔗 Привязка IP к именам пользователей</h1>
        <div class="nav-bar">
            <a href="/" class="btn">🏠 На главную</a>
            <a href="/admin/logs" class="btn">📜 Журнал действий</a>
            <a href="/admin/backups" class="btn">💾 Бэкапы</a>
            <a href="/changelog" class="btn">📋 Список изменений</a>
        </div>
        
        <form method="POST">
            <h3>Добавить / изменить привязку</h3>
            <input type="text" name="ip_address" placeholder="IP-адрес (например, 192.168.1.100)" required>
            <input type="text" name="username" placeholder="Имя пользователя" required>
            <input type="text" name="note" placeholder="Примечание (необязательно)">
            <label><input type="checkbox" name="access_allowed" value="1"> Доступ разрешён</label>
            <select name="role">
                <option value="viewer">Наблюдатель (только просмотр)</option>
                <option value="dispatcher">Диспетчер</option>
                <option value="supervisor">Супервизор (может редактировать данные)</option>
                <option value="admin">Администратор</option>
            </select>
            <button type="submit" class="btn" style="background:#27ae60;">Сохранить</button>
        </form>
        
        <h3>Текущие привязки</h3>
        <table>
            <thead>
                <tr>
                    <th>IP-адрес</th>
                    <th>Имя пользователя</th>
                    <th>Примечание</th>
                    <th>Доступ</th>
                    <th>Роль</th>
                    <th>Действия</th>
                </tr>
            </thead>
            <tbody>
    """
    for ip, username, note, access_allowed, role in rows:
        access_flag = "Да" if access_allowed else "Нет"
        html += f"""
                <tr>
                    <td>{ip}</td>
                    <td>{username}</td>
                    <td>{note or ''}</td>
                    <td>{access_flag}</td>
                    <td>{role}</td>
                    <td>
                        <form method="POST" style="margin:0; padding:0; display:inline;" onsubmit="return confirm('Удалить привязку для {ip}?')">
                            <input type="hidden" name="delete_ip" value="{ip}">
                            <button type="submit" class="delete-btn">❌ Удалить</button>
                        </form>
                    </td>
                </tr>
        """
    html += """
            </tbody>
        </table>
    </div>
    </body>
    </html>
    """
    return html

@app.route('/edit_wagon/<int:wagon_id>', methods=['POST'])
def edit_wagon_route(wagon_id):
    if request.user_role not in ('supervisor', 'admin'):
        return jsonify({"error": "Недостаточно прав"}), 403
    new_owner = request.form.get('owner') or None
    new_org = request.form.get('organization') or None
    new_note = request.form.get('note') or None
    new_arrival = request.form.get('arrival_time') or None
    new_global = request.form.get('departure_time') or None
    new_local = request.form.get('local_departure_time') or None
    success, msg = edit_wagon(wagon_id, new_owner, new_org, new_note, new_arrival, new_global, new_local)
    if success:
        return jsonify({"success": True, "message": msg})
    else:
        return jsonify({"success": False, "message": msg}), 400

@app.route('/edit_history/<int:history_id>', methods=['POST'])
def edit_history(history_id):
    if request.user_role not in ('supervisor', 'admin'):
        return jsonify({"error": "Недостаточно прав"}), 403
    
    new_timestamp_str = request.form.get('timestamp', '').strip()
    if not new_timestamp_str:
        return jsonify({"error": "Дата не указана"}), 400
    
    try:
        new_dt = parse_flexible_date(new_timestamp_str)
        if new_dt is None:
            raise ValueError
        new_timestamp = new_dt.strftime('%Y-%m-%d %H:%M:%S')
    except Exception as e:
        return jsonify({"error": f"Неверный формат даты: {e}"}), 400
    
    conn = get_conn()
    c = conn.cursor()
    
    c.execute("SELECT wagon_number, timestamp, action_type FROM movement_history WHERE id = ?", (history_id,))
    row = c.fetchone()
    if not row:
        conn.close()
        return jsonify({"error": "Запись не найдена"}), 404
    
    wagon_num, old_timestamp, action_type = row
    
    c.execute("""SELECT id, timestamp FROM movement_history 
                 WHERE wagon_number = ? AND id != ? 
                 ORDER BY timestamp ASC""", (wagon_num, history_id))
    all_events = c.fetchall()
    
    prev_ts = None
    next_ts = None
    found = False
    for ev_id, ev_ts in all_events:
        if ev_ts == old_timestamp and not found:
            found = True
            continue
        if not found:
            prev_ts = ev_ts
        else:
            next_ts = ev_ts
            break
    
    if prev_ts and new_timestamp <= prev_ts:
        conn.close()
        return jsonify({"error": f"Новая дата не может быть раньше предыдущего события ({prev_ts[:16]})"}), 400
    if next_ts and new_timestamp >= next_ts:
        conn.close()
        return jsonify({"error": f"Новая дата не может быть позже следующего события ({next_ts[:16]})"}), 400
    
    c.execute("UPDATE movement_history SET timestamp = ? WHERE id = ?", (new_timestamp, history_id))
    
    c.execute("SELECT MAX(timestamp) FROM movement_history WHERE wagon_number = ?", (wagon_num,))
    last_ts = c.fetchone()[0]
    if last_ts == new_timestamp:
        if action_type == 'added':
            c.execute("UPDATE wagons SET arrival_time = ? WHERE wagon_number = ? AND is_archived = 0", (new_timestamp, wagon_num))
        elif action_type == 'moved':
            c.execute("SELECT local_departure_time FROM wagons WHERE wagon_number = ? AND is_archived = 0", (wagon_num,))
            loc = c.fetchone()
            if loc and loc[0]:
                try:
                    old_loc_dt = datetime.strptime(loc[0], '%Y-%m-%d %H:%M:%S')
                    old_event_dt = datetime.strptime(old_timestamp, '%Y-%m-%d %H:%M:%S')
                    new_event_dt = datetime.strptime(new_timestamp, '%Y-%m-%d %H:%M:%S')
                    delta = old_loc_dt - old_event_dt
                    new_loc_dt = new_event_dt + delta
                    c.execute("UPDATE wagons SET local_departure_time = ? WHERE wagon_number = ? AND is_archived = 0",
                              (new_loc_dt.strftime('%Y-%m-%d %H:%M:%S'), wagon_num))
                except:
                    pass
    
    conn.commit()
    log_action('edit_history', wagon_number=wagon_num,
               details=f"Изменена дата события #{history_id} с {old_timestamp} на {new_timestamp}",
               old_value=old_timestamp, new_value=new_timestamp)
    conn.close()
    return jsonify({"success": True, "message": "Дата успешно обновлена"})

@app.route('/changelog')
def changelog():
    if request.user_role != 'admin':
        return "Доступ запрещён. Список изменений доступен только администраторам.", 403
    if not os.path.exists(CHANGELOG_PATH):
        return f"Файл CHANGELOG.txt не найден. Ожидаемый путь: {CHANGELOG_PATH}", 404
    with open(CHANGELOG_PATH, 'r', encoding='utf-8') as f:
        content = f.read()
    html = f"""
    <html>
    <head>
        <title>Список изменений</title>
        <style>
            body {{ font-family: monospace; padding: 20px; background: #f0f2f5; }}
            pre {{ background: white; padding: 20px; border-radius: 10px; overflow-x: auto; }}
            .btn {{ display: inline-block; padding: 5px 10px; background: #3498db; color: white; text-decoration: none; border-radius: 4px; margin-right: 5px; }}
            .nav-bar {{ margin-bottom: 15px; }}
        </style>
    </head>
    <body>
        <div class="nav-bar">
            <a href="/" class="btn">🏠 На главную</a>
            <a href="/admin/logs" class="btn">📜 Журнал действий</a>
            <a href="/admin/ip_users" class="btn">🔗 Привязка IP</a>
            <a href="/admin/backups" class="btn">📦 Резервные копии</a>
        </div>
        <h1>📋 Список изменений</h1>
        <pre>{content}</pre>
    </body>
    </html>
    """
    return html

@app.route('/help')
def help_page():
    return render_template_string(HELP_TEMPLATE, server_ip=SERVER_IP or "IP_вашего_сервера", version=APP_VERSION)

@app.route('/')
def index():
    tracks, move_list = get_dashboard_data()
    is_admin = (request.user_role == 'admin')
    return render_template_string(HTML_TEMPLATE, tracks=tracks, move_list=move_list, total_wagons=len(move_list),
                                  add_form_data=None, move_form_data=None, request=request, is_admin=is_admin,
                                  version=APP_VERSION)

@app.route('/add', methods=['POST'])
def add_wagon():
    number = request.form.get('number', '').strip()
    owner = request.form.get('owner', '').strip()
    org = request.form.get('organization', '').strip()
    note = request.form.get('note', '')
    track_id_str = request.form.get('track_id', '')
    cycle_days = request.form.get('cycle_days', '0')
    cycle_hours = request.form.get('cycle_hours', '0')
    cycle_mins = request.form.get('cycle_mins', '0')
    start_date = request.form.get('start_date', '').strip()
    start_time = request.form.get('start_time', '').strip()
    add_form_data = {
        'number': number,
        'owner': owner,
        'org': org,
        'note': note,
        'cycle_days': cycle_days,
        'cycle_hours': cycle_hours,
        'cycle_mins': cycle_mins,
        'start_date': start_date,
        'start_time': start_time,
        'track_id': track_id_str
    }
    if not number or not owner or not org or not track_id_str:
        flash("Заполните все поля!", 'error')
        tracks, move_list = get_dashboard_data()
        is_admin = (request.user_role == 'admin')
        return render_template_string(HTML_TEMPLATE, tracks=tracks, move_list=move_list, total_wagons=len(move_list),
                                      add_form_data=add_form_data, move_form_data=None, request=request, is_admin=is_admin)
    try:
        track_id = int(track_id_str)
    except ValueError:
        flash("Ошибка: Неверный ID пути.", 'error')
        tracks, move_list = get_dashboard_data()
        is_admin = (request.user_role == 'admin')
        return render_template_string(HTML_TEMPLATE, tracks=tracks, move_list=move_list, total_wagons=len(move_list),
                                      add_form_data=add_form_data, move_form_data=None, request=request, is_admin=is_admin)
    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT id, status, is_archived FROM wagons WHERE wagon_number = ?", (number,))
    existing = c.fetchone()
    try:
        days = int(cycle_days) if cycle_days else 0
        hours = int(cycle_hours) if cycle_hours else 0
        mins = int(cycle_mins) if cycle_mins else 0
    except ValueError:
        days, hours, mins = 0, 0, 0
    total_mins = (days * 24 * 60) + (hours * 60) + mins
    if (start_date and not start_time) or (start_time and not start_date):
        flash("Ошибка: Если вы указываете дату, нужно указать и время, и наоборот. Либо оставьте оба поля пустыми.", 'error')
        tracks, move_list = get_dashboard_data()
        is_admin = (request.user_role == 'admin')
        return render_template_string(HTML_TEMPLATE, tracks=tracks, move_list=move_list, total_wagons=len(move_list),
                                      add_form_data=add_form_data, move_form_data=None, request=request, is_admin=is_admin)
    manual_start = None
    if start_date and start_time:
        manual_start = f"{start_date} {start_time}"
        try:
            start_dt = datetime.strptime(manual_start, '%Y-%m-%d %H:%M')
            arrival_time = f"{start_date} {start_time}:00"
        except:
            arrival_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            flash("Неверный формат даты/времени, использовано текущее время", 'warning')
    else:
        arrival_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    global_dep = None
    if total_mins > 0:
        if manual_start:
            try:
                start_dt = datetime.strptime(manual_start, '%Y-%m-%d %H:%M')
                global_dep = (start_dt + timedelta(minutes=total_mins)).strftime('%Y-%m-%d %H:%M:%S')
            except:
                global_dep = (datetime.now() + timedelta(minutes=total_mins)).strftime('%Y-%m-%d %H:%M:%S')
        else:
            global_dep = (datetime.now() + timedelta(minutes=total_mins)).strftime('%Y-%m-%d %H:%M:%S')
    if existing:
        w_id, w_status, w_archived = existing
        if w_archived == 1:
            compact_track(track_id)
            pos = find_slot_on_track(track_id, 10)[1]
            c.execute("""UPDATE wagons SET status = 'assigned', owner = ?, organization = ?, cargo_type = ?, track_id = ?, start_pos = ?, arrival_time = ?, departure_time = ?, local_departure_time = NULL, visit_count = 0, is_archived = 0 WHERE id = ?""", 
                      (owner, org, note, track_id, float(pos), arrival_time, global_dep, w_id))
            conn.commit()
            conn.close()
            log_movement(number, 'added', None, None, f"Восстановлен из архива. ТК: {owner}, Орг: {org}", arrival_time)
            log_action('add', wagon_number=number, details=f"Восстановлен из архива на путь {track_id_str}")
            flash(f"✅ Вагон {number} восстановлен с временем прибытия {arrival_time[:16]}.", 'success')
            return redirect(url_for('index'))
        elif w_status != 'departed':
            conn.close()
            flash(f"⚠️ Вагон '{number}' уже на путях!", 'error')
            tracks, move_list = get_dashboard_data()
            is_admin = (request.user_role == 'admin')
            return render_template_string(HTML_TEMPLATE, tracks=tracks, move_list=move_list, total_wagons=len(move_list),
                                          add_form_data=add_form_data, move_form_data=None, request=request, is_admin=is_admin)
    compact_track(track_id)
    pos = find_slot_on_track(track_id, 10)[1]
    try:
        c.execute("""INSERT INTO wagons (wagon_number, length, cargo_type, owner, organization, track_id, start_pos, arrival_time, departure_time, local_departure_time, visit_count, is_archived) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 0, 0)""", 
                  (number, 10.0, note, owner, org, track_id, float(pos), arrival_time, global_dep, None))
        conn.commit()
        t_name = c.execute("SELECT name FROM tracks WHERE id=?", (track_id,)).fetchone()[0]
        conn.close()
        log_movement(number, 'added', None, t_name, f"ТК: {owner}, Орг: {org}", arrival_time)
        log_action('add', wagon_number=number, details=f"Добавлен на путь {t_name}. ТК: {owner}, Орг: {org}")
        msg = f"✅ Вагон {number} добавлен с временем прибытия {arrival_time[:16]}."
        if total_mins > 0:
            msg += f" Срок: {days}д {hours}ч {mins}мин"
        if global_dep:
            msg += f" (до {global_dep[:16]})"
        flash(msg, 'success')
    except sqlite3.IntegrityError:
        conn.close()
        flash(f"⚠️ Вагон '{number}' уже существует.", 'error')
        tracks, move_list = get_dashboard_data()
        is_admin = (request.user_role == 'admin')
        return render_template_string(HTML_TEMPLATE, tracks=tracks, move_list=move_list, total_wagons=len(move_list),
                                      add_form_data=add_form_data, move_form_data=None, request=request, is_admin=is_admin)
    return redirect(url_for('index'))

@app.route('/move', methods=['POST'])
def move_action():
    wagon_id = request.form.get('wagon_id', '')
    new_track_id_str = request.form.get('new_track_id', '')
    local_days = request.form.get('local_days', '0')
    local_hours = request.form.get('local_hours', '0')
    local_mins = request.form.get('local_mins', '0')
    start_date = request.form.get('start_date', '').strip()
    start_time = request.form.get('start_time', '').strip()
    note = request.form.get('note', '')
    move_form_data = {
        'wagon_id': wagon_id,
        'new_track_id': new_track_id_str,
        'local_days': local_days,
        'local_hours': local_hours,
        'local_mins': local_mins,
        'start_date': start_date,
        'start_time': start_time,
        'note': note
    }
    if not wagon_id or not new_track_id_str:
        flash("Выберите вагон и путь назначения!", 'error')
        tracks, move_list = get_dashboard_data()
        is_admin = (request.user_role == 'admin')
        return render_template_string(HTML_TEMPLATE, tracks=tracks, move_list=move_list, total_wagons=len(move_list),
                                      add_form_data=None, move_form_data=move_form_data, request=request, is_admin=is_admin)
    try:
        new_track_id = int(new_track_id_str)
    except ValueError:
        flash("Ошибка: Неверный ID пути.", 'error')
        tracks, move_list = get_dashboard_data()
        is_admin = (request.user_role == 'admin')
        return render_template_string(HTML_TEMPLATE, tracks=tracks, move_list=move_list, total_wagons=len(move_list),
                                      add_form_data=None, move_form_data=move_form_data, request=request, is_admin=is_admin)
    try:
        l_days = int(local_days) if local_days else 0
        l_hours = int(local_hours) if local_hours else 0
        l_mins = int(local_mins) if local_mins else 0
    except ValueError:
        l_days, l_hours, l_mins = 0, 0, 0
    if (start_date and not start_time) or (start_time and not start_date):
        flash("Ошибка: Если вы указываете дату, нужно указать и время, и наоборот. Либо оставьте оба поля пустыми.", 'error')
        tracks, move_list = get_dashboard_data()
        is_admin = (request.user_role == 'admin')
        return render_template_string(HTML_TEMPLATE, tracks=tracks, move_list=move_list, total_wagons=len(move_list),
                                      add_form_data=None, move_form_data=move_form_data, request=request, is_admin=is_admin)
    manual_start = None
    if start_date and start_time:
        manual_start = f"{start_date} {start_time}"
    success, msg = move_wagon(wagon_id, new_track_id, l_days, l_hours, l_mins, manual_start, note)
    if success:
        flash(msg, 'success')
        return redirect(url_for('index'))
    else:
        flash(msg, 'error')
        tracks, move_list = get_dashboard_data()
        is_admin = (request.user_role == 'admin')
        return render_template_string(HTML_TEMPLATE, tracks=tracks, move_list=move_list, total_wagons=len(move_list),
                                      add_form_data=None, move_form_data=move_form_data, request=request, is_admin=is_admin)

@app.route('/depart/<int:wagon_id>', methods=['POST'])
def depart_action(wagon_id):
    if depart_wagon(wagon_id): 
        flash("✅ Вагон убран в архив.", 'success')
    else: 
        flash("⚠️ Ошибка при удалении.", 'error')
    return redirect(url_for('index'))

def create_tray_icon():
    icon_path = None
    for ext in ['.png', '.ico']:
        test_path = os.path.join(BASE_DIR, f'icon{ext}')
        if os.path.exists(test_path):
            icon_path = test_path
            break
    if icon_path:
        try:
            img = Image.open(icon_path)
            img = img.resize((64, 64), Image.Resampling.LANCZOS if hasattr(Image, 'Resampling') else Image.ANTIALIAS)
        except:
            img = None
    if not icon_path or img is None:
        img = Image.new('RGB', (64, 64), color='#2c3e50')
        draw = ImageDraw.Draw(img)
        draw.rectangle([(0, 50), (64, 58)], fill='#7f8c8d')
        draw.rectangle([(0, 55), (64, 60)], fill='#95a5a6')
        draw.rectangle([(10, 30), (54, 48)], fill='#e74c3c', outline='white', width=1)
        draw.ellipse([(16, 45), (24, 53)], fill='#2c3e50', outline='white', width=1)
        draw.ellipse([(40, 45), (48, 53)], fill='#2c3e50', outline='white', width=1)
        draw.rectangle([(14, 34), (22, 42)], fill='#3498db')
        draw.rectangle([(26, 34), (34, 42)], fill='#3498db')
        draw.rectangle([(38, 34), (46, 42)], fill='#3498db')
        draw.rectangle([(12, 28), (52, 32)], fill='#c0392b')
        draw.ellipse([(48, 20), (56, 28)], fill='#bdc3c7')
        draw.ellipse([(52, 14), (60, 22)], fill='#bdc3c7')
    def on_open(icon, item):
        webbrowser.open('http://127.0.0.1:5000')
    def on_quit(icon, item):
        icon.stop()
        os._exit(0)
    menu = pystray.Menu(
        pystray.MenuItem("🚂 Открыть", on_open),
        pystray.MenuItem("❌ Выход", on_quit)
    )
    icon = pystray.Icon("railway_dispatcher", img, "ЖД Диспетчерская", menu)
    icon.run()

if __name__ == '__main__':
    init_db()
    clean_action_log()
    last_auto = get_last_auto_backup_time()
    if last_auto is None or (datetime.now() - last_auto) > timedelta(days=1):
        print("📦 Создание автоматического бэкапа при запуске (прошло более суток)...")
        create_auto_backup()
    schedule_daily_backup()
    print("=" * 50)
    print("🚂 ЖД Диспетчерская запущена (версия 2.4.8b)")
    print("=" * 50)
    print(f"📁 База данных: {DB_NAME}")
    print(f"🌐 Локальный адрес: http://127.0.0.1:5000")
    try:
        import socket
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        local_ip = s.getsockname()[0]
        s.close()
        print(f"🌐 Сетевой адрес: http://{local_ip}:5000")
        SERVER_IP = local_ip
    except:
        SERVER_IP = "IP_сервера"
        pass
    print("=" * 50)
    print("Для остановки нажмите Ctrl+C или используйте иконку в трее")
    print("=" * 50)
    server_thread = threading.Thread(target=lambda: app.run(host='0.0.0.0', port=5000, debug=False, use_reloader=False), daemon=True)
    server_thread.start()
    if HAS_TRAY:
        create_tray_icon()
    else:
        print("Иконка в трее не доступна. Для выхода нажмите Ctrl+C")
        server_thread.join()